import json
import math
import asyncio
import ssl
import io
import base64
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg') # CRITICAL: Headless Mode for Ubuntu!
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.patches import Rectangle, Ellipse, Polygon
import matplotlib.gridspec as gridspec
from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
import paho.mqtt.client as mqtt

app = FastAPI()

# Allow frontend web dashboard to connect
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ==========================================
# 1. CLOUD CONFIGURATION & MEMORY VAULT
# ==========================================
AWS_ENDPOINT = ""  # <--- Keep your actual endpoint here!
MQTT_TOPIC_SUB = ""
MQTT_TOPIC_PUB = ""     # <--- NEW: The Downlink Channel

# --- THE CLOUD DATA HISTORIAN (For Excel) ---
history_dist = []
history_rms = []
history_worst = []
history_sensor = {0:[], 1:[], 2:[], 3:[], 4:[]}
history_details = {i: {'dL':[], 'da':[], 'db':[]} for i in range(5)}

# ==========================================
# 2. THE MATHEMATICAL VAULT
# ==========================================
SIGMA_FACTOR_A = math.sqrt(-2 * math.log(0.94))
CURRENT_SIGMA = 1.50
INV_SIGMA = 1.0 / CURRENT_SIGMA  # <--- PREMIUM FIX: Pre-computed inverse for O(1) multiplication
CURRENT_CALIB_MODE = "MASTER"

# In a full app, these come from a database. For now, we initialize a safe baseline.
vault_data = {i: {'L': 50.0, 'a': 0.0, 'b': 0.0} for i in range(5)}
active_sensors = [True, True, True, True, True]
_cmc_weight_cache = {}

def calc_cmc(L_std, a_std, b_std, dL, da, db, l=2.0, c=0.4):
    C_std = math.sqrt((a_std * a_std) + (b_std * b_std))
    a_samp = a_std + da
    b_samp = b_std + db
    C_samp = math.sqrt((a_samp * a_samp) + (b_samp * b_samp))
    dC = C_samp - C_std
    dH_sq = max(0.0, ((da * da) + (db * db)) - (dC * dC)) 

    cache_key = (L_std, a_std, b_std)
    global _cmc_weight_cache
    if cache_key in _cmc_weight_cache:
        inv_L_weight, inv_C_weight, inv_H_sq_weight, S_C, S_H, S_L = _cmc_weight_cache[cache_key]
    else:
        S_L = 0.511 if L_std < 16.0 else (0.040975 * L_std) / (1.0 + 0.01765 * L_std)
        S_C = ((0.0638 * C_std) / (1.0 + 0.0131 * C_std)) + 0.638
        if C_std < 0.0001: 
            S_H = S_C 
        else:
            h_std = math.degrees(math.atan2(b_std, a_std))
            if h_std < 0: h_std += 360.0
            C_std_sq = C_std * C_std
            C_std_4 = C_std_sq * C_std_sq
            F = math.sqrt(C_std_4 / (C_std_4 + 1900.0))
            if 164.0 <= h_std <= 345.0:
                T = 0.56 + abs(0.2 * math.cos(math.radians(h_std + 168.0)))
            else:
                T = 0.36 + abs(0.4 * math.cos(math.radians(h_std + 35.0)))
            S_H = S_C * (F * T + 1.0 - F)

        inv_L_weight = 1.0 / (l * S_L)
        inv_C_weight = 1.0 / (c * S_C)
        inv_H_sq_weight = 1.0 / (S_H * S_H)
        _cmc_weight_cache[cache_key] = (inv_L_weight, inv_C_weight, inv_H_sq_weight, S_C, S_H, S_L)

    term_L = dL * inv_L_weight
    term_C = dC * inv_C_weight
    term_H_sq = dH_sq * inv_H_sq_weight
    de_cmc = math.sqrt((term_L * term_L) + (term_C * term_C) + term_H_sq)
    return de_cmc, S_C, S_H, S_L

def generate_cloud_excel():
    SIGMA_FACTOR_B = math.sqrt(-2 * math.log(0.89))
    current_sigma = CURRENT_SIGMA
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    default_name = f"TMDM_Cloud_Report_{timestamp}.xlsx"

    total_points = len(history_dist)
    if total_points == 0:
        raise ValueError("No data to export")

    dists = list(history_dist)
    s_data_list = [list(history_sensor[i]) for i in range(5)]
    rms_des = list(history_rms)
    worst_des = list(history_worst)
    det_data = {i: {k: list(history_details[i][k]) for k in ['dL', 'da', 'db']} for i in range(5)}

    de_a = current_sigma * SIGMA_FACTOR_A
    de_b = current_sigma * SIGMA_FACTOR_B

    grade_a_count = sum(1 for rms in rms_des if rms <= de_a)
    grade_b_count = sum(1 for rms in rms_des if de_a < rms <= de_b)
    grade_c_count = sum(1 for rms in rms_des if rms > de_b)
    
    # --- PREMIUM FIX: Single Inverse Multiplier for the entire Excel Sheet ---
    inv_total_pct = (1.0 / total_points) * 100.0 if total_points > 0 else 0.0
    
    yield_a = grade_a_count * inv_total_pct
    yield_b = grade_b_count * inv_total_pct
    yield_c = grade_c_count * inv_total_pct

    if len(worst_des) > 0:
        max_worst_de = max(worst_des)
        worst_index = worst_des.index(max_worst_de)
        max_worst_dist = dists[worst_index]
        max_worst_sensor_id = "Unknown"
        for s in range(5):
            if active_sensors[s] and s_data_list[s][worst_index] == max_worst_de:
                max_worst_sensor_id = f"#{s}"
                break
    else:
        max_worst_de = 0.0; max_worst_dist = 0.0; max_worst_sensor_id = "N/A"

    sensor_stats = {}
    for i in range(5):
        if active_sensors[i]:
            s_data = s_data_list[i]
            s_count_a = sum(1 for val in s_data if val <= de_a)
            s_count_b = sum(1 for val in s_data if de_a < val <= de_b)
            s_count_c = sum(1 for val in s_data if val > de_b)
            s_max = max(s_data) if len(s_data) > 0 else 0.0
            s_max_dist = dists[s_data.index(s_max)] if len(s_data) > 0 else 0.0
            # --- PREMIUM FIX: O(1) Multiplication using the inv_total_pct cache! ---
            # (Note: We can also drop the 'if total_points > 0' check here because 
            # inv_total_pct safely defaults to 0.0 at the top of the function if points are 0!)
            sensor_stats[i] = {
                'A': round(s_count_a * inv_total_pct, 1),
                'B': round(s_count_b * inv_total_pct, 1),
                'C': round(s_count_c * inv_total_pct, 1),
                'Max': round(s_max, 2), 'MaxDist': round(s_max_dist, 3)
            }
        else:
            sensor_stats[i] = {'A': "N/A", 'B': "N/A", 'C': "N/A", 'Max': "N/A", 'MaxDist': "N/A"}

    if len(worst_des) > 0:
        sorted_worst = sorted(worst_des)
        keep_count = max(1, int(len(sorted_worst) * 0.99))
        bottom_99_array = sorted_worst[:keep_count]
        true_avg_de = sum(bottom_99_array) / len(bottom_99_array) if len(bottom_99_array) > 0 else 0.0
        
        ratio_99 = true_avg_de / current_sigma
        p99_var = 100.0 - (100.0 * math.exp(-0.5 * (ratio_99 * ratio_99)))
            
        p100_avg_de = sum(worst_des) / len(worst_des)
        ratio_100 = p100_avg_de / current_sigma
        p100_var = 100.0 - (100.0 * math.exp(-0.5 * (ratio_100 * ratio_100)))
        outlier_impact = ((p100_avg_de - true_avg_de) / true_avg_de) * 100.0 if true_avg_de > 0 else 0.0
    else:
        p99_var = 0.0; p100_var = 0.0; outlier_impact = 0.0

    if outlier_impact <= 2.0: oir_warning = "Excellent: Roll is highly consistent. No severe isolated stains detected."
    elif outlier_impact <= 5.0: oir_warning = "Notice: Moderate variance. A few noticeable but non-critical spikes exist."
    else: oir_warning = "Warning! The base fabric is fine, but there is a massive, severe stain somewhere on this roll that is ruining the average!"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shade Inspection Data"
    
    ws.append(["TMDM Continuous Shade Inspection Final Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1') 
    ws.append(["Export Date/Time:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Strictness Tolerance (Sigma):", round(current_sigma, 2)])
    ws.append(["Passed Target (ΔE):", round(de_a, 2)])
    ws.append(["Total Fabric Scanned:", f"{dists[-1] if len(dists)>0 else 0.0:.3f} Yards"])
    ws.append(["Session Duration:", "Tracked on Client Web Dashboard"])
    ws.append([])
    
    ws.append(["--- OVERALL ROLL QUALITY (MASTER RMS) ---"])
    r1 = ws.max_row
    ws[f"A{r1}"].font = Font(bold=True, color="00529B") 
    ws.merge_cells(f"A{r1}:E{r1}")
    
    for label, val in [("Final Yield (Grade A) %:", yield_a), ("Final Yield (Grade B) %:", yield_b), ("Final Yield (Grade C) %:", yield_c), ("Average Worst Variation (P99) %:", p99_var), ("Absolute Worst Variation (P100) %:", p100_var)]:
        ws.append([label, round(val, 1)])
        ws[f"B{ws.max_row}"].number_format = '0.0"%"'
    
    ws.append(["Outlier Impact Ratio (Spike Intensity) %:", round(outlier_impact, 1), oir_warning])
    ws[f"B{ws.max_row}"].number_format = '0.0"%"'
    ws.append(["Absolute Worst Spot Detected:", f"{max_worst_de:.2f} ΔE (Occurred at {max_worst_dist:.3f} Yards) by Sensor {max_worst_sensor_id}"])
    ws.append([])

    ws.append(["--- OIR (OUTLIER IMPACT RATIO) LEGEND ---"])
    r2 = ws.max_row
    ws[f"A{r2}"].font = Font(bold=True, color="00529B")
    ws.merge_cells(f"A{r2}:B{r2}")
    ws.append(["0% - 2%", "Highly Consistent: P100 and P99 are nearly identical. No sudden oil stains or color drops."])
    ws.append(["2.1% - 5%", "Moderate Variance: A few noticeable but non-critical spikes exist on the roll."])
    ws.append(["> 5%", "Severe Defect Alert: The base fabric is fine, but a massive, severe stain is ruining the average!"])
    ws.append([])

    ws.append(["--- SQI VARIATION & GRADING LEGEND ---"])
    r_sqi = ws.max_row
    ws[f"A{r_sqi}"].font = Font(bold=True, color="00529B")
    ws.merge_cells(f"A{r_sqi}:B{r_sqi}")
    ws.append(["Grade A (<= 6.0%)", "Premium Quality: Excellent shade consistency. Meets strictest buyer tolerances."])
    ws.append(["Grade B (6.1% - 11.0%)", "Standard Quality: Acceptable commercial match. Noticeable but within passable limits."])
    ws.append(["Grade C (> 11.0%)", "Rejected: Severe shade band or heavy staining. Fails quality assurance."])
    ws.append([])

    ws.append(["--- INDIVIDUAL SENSOR YIELD DASHBOARD ---"])
    r3 = ws.max_row
    ws[f"A{r3}"].font = Font(bold=True, color="00529B")
    ws.merge_cells(f"A{r3}:L{r3}") 
    
    # --- PREMIUM FIX: Dynamic Headers for Manual QTX Mode ---
    if CURRENT_CALIB_MODE == "MANUAL":
        ws.append(["Sensor Node", "Status", "Manual L1", "Manual a1", "Manual b1", "Manual C1*", "Manual h1°", f"Grade A (<={de_a:.2f}) %", f"Grade B (<={de_b:.2f}) %", f"Grade C (>{de_b:.2f}) %", "Max Spike (ΔE CMC)", "Spike Distance (Yards)"])
    else:
        ws.append(["Sensor Node", "Status", "Master L*", "Master a*", "Master b*", "Master C*", "Master h°", f"Grade A (<={de_a:.2f}) %", f"Grade B (<={de_b:.2f}) %", f"Grade C (>{de_b:.2f}) %", "Max Spike (ΔE CMC)", "Spike Distance (Yards)"])
    dash_header_row = ws.max_row
    for col_idx in range(1, 13):
        ws.cell(row=dash_header_row, column=col_idx).font = Font(bold=True)
        ws.cell(row=dash_header_row, column=col_idx).alignment = Alignment(horizontal="center")
        
    for i in range(5):
        m_L = vault_data[i].get('L', 0.0); m_a = vault_data[i].get('a', 0.0); m_b = vault_data[i].get('b', 0.0)
        status = "ACTIVE" if active_sensors[i] else "DISABLED"
        if active_sensors[i]:
            m_C = math.sqrt((m_a * m_a) + (m_b * m_b))
            m_h_deg = math.degrees(math.atan2(m_b, m_a))
            if m_h_deg < 0: m_h_deg += 360.0
        else:
            m_C = 0.0; m_h_deg = 0.0
        
        ws.append([f"Sensor {i}", status, 
            round(m_L, 2) if active_sensors[i] else "N/A", round(m_a, 2) if active_sensors[i] else "N/A", round(m_b, 2) if active_sensors[i] else "N/A", round(m_C, 2) if active_sensors[i] else "N/A", round(m_h_deg, 2) if active_sensors[i] else "N/A", 
            sensor_stats[i]['A'], sensor_stats[i]['B'], sensor_stats[i]['C'], sensor_stats[i]['Max'], sensor_stats[i]['MaxDist']])
        for col_idx in range(2, 13):
            ws.cell(row=ws.max_row, column=col_idx).alignment = Alignment(horizontal="center")
    ws.append([])

    center_alignment = Alignment(horizontal="center", vertical="center")
    fill_light_blue = PatternFill(start_color="33CCFF", end_color="33CCFF", fill_type="solid")
    fill_title = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid") 
    header_rows_ws1 = [dash_header_row]
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.startswith("---"): header_rows_ws1.append(r)
    
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if cell.row == 1: cell.fill = fill_title
            elif cell.row in header_rows_ws1: cell.fill = fill_light_blue
            if type(cell).__name__ != 'MergedCell':
                if col_letter is None: col_letter = cell.column_letter
                if cell.value is not None:
                    if cell.row > 1: max_length = max(max_length, len(str(cell.value)))
                    cell.alignment = center_alignment
                    if cell.row == 1: cell.font = Font(bold=True, size=14, color="7030A0")
                    else:
                        is_orange = False
                        if cell.column == 1 and isinstance(cell.value, str):
                            text = cell.value
                            if text.endswith(":") or text.startswith("0%") or text.startswith("2.1%") or text.startswith("> 5%") or text.startswith("Grade A (") or text.startswith("Grade B (") or text.startswith("Grade C ("): is_orange = True
                        cell.font = Font(bold=True, color="FF8C00") if is_orange else (Font(bold=True, color="7030A0") if cell.font and cell.font.bold else Font(color="7030A0"))
        if col_letter and max_length > 0: ws.column_dimensions[col_letter].width = max_length + 2

    ws2 = wb.create_sheet("Observation Data")
    from openpyxl.styles import Border, Side
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- RESTORED: PREMIUM DUAL-TRIGGER SNAPSHOT ENGINE ---
    next_interval_target = 2.0
    
    # =====================================================================
    # --- PREMIUM FIX: PER-SENSOR INDEPENDENT TRACKERS ---
    # Tracks the absolute worst moment for EVERY sensor individually!
    # =====================================================================
    running_max_de = [-1.0] * 5
    running_target_idx = [0] * 5
    
    # --- PREMIUM FIX: BLOCK-LEVEL TRACKERS ---
    block_state_failing = False
    interval_start_dist = 0.0  # <--- PREMIUM FIX: Tracks the Entering Distance
    
    de_a_str = f"{de_a:.2f}"

    master_cache = {}
    for s in range(5):
        if active_sensors[s]:
            m_L_cache = vault_data[s].get('L', 0.0); m_a_cache = vault_data[s].get('a', 0.0); m_b_cache = vault_data[s].get('b', 0.0)
            m_C_cache = math.sqrt((m_a_cache * m_a_cache) + (m_b_cache * m_b_cache))
            m_h_deg_cache = math.degrees(math.atan2(m_b_cache, m_a_cache))
            if m_h_deg_cache < 0: m_h_deg_cache += 360.0
            master_cache[s] = {'m_a': m_a_cache, 'm_b': m_b_cache, 'm_C': m_C_cache, 'm_h_deg': m_h_deg_cache}

    # =====================================================================
    # --- PREMIUM FIX: THE 300-DPI "LEAN & MEAN" O(1) ENGINE ---
    # Eradicates the 54MB crash by restoring tight crop boundaries while 
    # keeping the lightning-fast O(1) background and legend bypass!
    # =====================================================================
    report_fig = Figure(figsize=(5.0, 2.5), dpi=300) 
    report_canvas = FigureCanvasAgg(report_fig)
    
    gs = gridspec.GridSpec(1, 2, width_ratios=[3.5, 1.5], wspace=0.35)
    ax_wheel = report_fig.add_subplot(gs[0])
    ax_bar = report_fig.add_subplot(gs[1])
    
    # =====================================================================
    # --- PREMIUM FIX: SENSOR 4 MASTER ANCHOR PRIORITY ---
    # Bases the visual limits on Sensor 4 if active. Otherwise, falls back gracefully.
    # =====================================================================
    anchor_s = 4 if active_sensors[4] else next((s for s in range(5) if active_sensors[s]), 0)
    
    m_L = vault_data[anchor_s].get('L', 0.0); m_a = master_cache[anchor_s]['m_a']; m_b = master_cache[anchor_s]['m_b']
    
    if 'm_h_rad' not in master_cache[anchor_s]:
        master_cache[anchor_s]['m_h_rad'] = math.atan2(m_b, m_a)
    master_hue_rad = master_cache[anchor_s]['m_h_rad']; master_hue_deg = master_cache[anchor_s]['m_h_deg']
    _, S_C, S_H, S_L = calc_cmc(m_L, m_a, m_b, 0.0, 0.0, 0.0)

    cmc_radius = de_a * S_C
    l_bound = de_a * (2.0 * S_L) 

    # =====================================================================
    # --- PREMIUM FIX: THE GLOBAL SMART ZOOM ENGINE ---
    # Scans the entire roll's history ONCE to calculate the perfect O(1) zoom!
    # Restores the beautiful large ellipse while guaranteeing no outliers fall off the chart.
    # =====================================================================
    global_max_ab = 0.5
    global_max_L = 0.5
    for s in range(5):
        if active_sensors[s]:
            if det_data[s]['da']: global_max_ab = max(global_max_ab, max(map(abs, det_data[s]['da'])))
            if det_data[s]['db']: global_max_ab = max(global_max_ab, max(map(abs, det_data[s]['db'])))
            if det_data[s]['dL']: global_max_L = max(global_max_L, max(map(abs, det_data[s]['dL'])))
            
    limit = math.ceil(global_max_ab * 1.3 * 10.0) / 10.0
    if limit < cmc_radius * 1.5: limit = cmc_radius * 1.5
    
    limit_L = math.ceil(global_max_L * 1.3 * 10.0) / 10.0
    if limit_L < l_bound * 1.5: limit_L = l_bound * 1.5

    ax_wheel.set_xlim(-limit, limit); ax_wheel.set_ylim(-limit, limit); ax_wheel.set_aspect('equal', adjustable='box')
    ax_wheel.add_patch(Polygon([(-limit, limit), (0,0), (limit, limit)], color='#FFFCA6', alpha=0.5))
    ax_wheel.add_patch(Polygon([(-limit, -limit), (0,0), (limit, -limit)], color='#E6E6FF', alpha=0.6))
    ax_wheel.add_patch(Polygon([(limit, limit), (0,0), (limit, -limit)], color='#FFE6E6', alpha=0.5))
    ax_wheel.add_patch(Polygon([(-limit, limit), (0,0), (-limit, -limit)], color='#E6FFE6', alpha=0.5))
    ax_wheel.axhline(0, color='black', linewidth=1.0, alpha=0.6); ax_wheel.axvline(0, color='black', linewidth=1.0, alpha=0.6)

    num_ticks_w = 7; step_w = (limit * 2) / (num_ticks_w - 1); tick_len = limit * 0.03
    for tick_idx in range(num_ticks_w):
        t_val = -limit + (tick_idx * step_w)
        ax_wheel.plot([t_val, t_val], [-tick_len, tick_len], color='black', lw=0.8, alpha=0.6)
        ax_wheel.plot([-tick_len, tick_len], [t_val, t_val], color='black', lw=0.8, alpha=0.6)
        if abs(t_val) > 0.01:
            ax_wheel.text(t_val, -limit*0.06, f"{t_val:.1f}", fontsize=5, ha='center', va='top', alpha=0.8)
            ax_wheel.text(-limit*0.06, t_val, f"{t_val:.1f}", fontsize=5, ha='right', va='center', alpha=0.8)

    r_C = de_a * S_C; end_x = math.cos(master_hue_rad) * r_C; end_y = math.sin(master_hue_rad) * r_C
    r_H = de_a * S_H; perp_rad = master_hue_rad + (math.pi * 0.5); p_x = math.cos(perp_rad) * r_H; p_y = math.sin(perp_rad) * r_H
    ax_wheel.plot([-end_x, end_x], [-end_y, end_y], color='black', linestyle='--', linewidth=1.0, alpha=0.7)
    ax_wheel.plot(end_x, end_y, marker=(3, 0, master_hue_deg - 90), markersize=6, color='black', alpha=0.8)
    ax_wheel.plot([-p_x, p_x], [-p_y, p_y], color='black', linestyle='--', linewidth=1.0, alpha=0.7)

    e_width = (de_a * S_C) * 2; e_height = (de_a * S_H) * 2
    ax_wheel.add_patch(Ellipse((0,0), width=e_width, height=e_height, angle=master_hue_deg, edgecolor='#FF0000', facecolor='none', linestyle='-', linewidth=1.2, alpha=0.8))
    ax_wheel.add_patch(Ellipse((0,0), width=e_width * 1.2, height=e_height * 1.2, angle=master_hue_deg, edgecolor='gray', facecolor='none', linestyle='--', linewidth=0.8, alpha=0.7))

    ax_wheel.text(limit*0.95, 0.05 * limit, '+a*', fontsize=7, ha='right'); ax_wheel.text(-limit*0.95, 0.05 * limit, '-a*', fontsize=7, ha='left')
    ax_wheel.text(0.05 * limit, limit*0.85, '+b*', fontsize=7, ha='left'); ax_wheel.text(0.05 * limit, -limit*0.9, '-b*', fontsize=7, ha='left')
    ax_wheel.plot(0, 0, marker='+', color='black', markersize=8, mew=1.2)

    ax_bar.set_xlim(-0.6, 0.6); ax_bar.set_ylim(-limit_L, limit_L); margin = l_bound * 0.1 
    ax_bar.add_patch(Rectangle((-0.25, -l_bound), 0.5, l_bound * 2, color='#A8E6A8', alpha=0.8)) 
    ax_bar.add_patch(Rectangle((-0.25, l_bound), 0.5, margin, color='#FFFFA6', alpha=0.9)) 
    ax_bar.add_patch(Rectangle((-0.25, -l_bound - margin), 0.5, margin, color='#FFFFA6', alpha=0.9)) 
    ax_bar.axvline(0, color='black', linewidth=1.2, alpha=0.4); ax_bar.axhline(0, color='black', linewidth=1.5); ax_bar.plot(0, 0, marker='+', color='black', markersize=8, mew=1.2)

    num_ticks = 9; step = (limit_L * 2) / (num_ticks - 1)
    for tick_idx in range(num_ticks):
        t_val = -limit_L + (tick_idx * step)
        ax_bar.axhline(t_val, color='black', linewidth=0.8, alpha=0.5, xmin=0.35, xmax=0.65)
        if abs(t_val) > 0.01: ax_bar.text(-0.35, t_val, f"{t_val:.1f}", fontsize=6, va='center', ha='right')

    ax_bar.text(0, limit_L * 1.05, '+L* (Lighter)', fontsize=7, ha='center', fontweight='bold')
    ax_bar.text(0, -limit_L * 1.15, '-L* (Darker)', fontsize=7, ha='center', fontweight='bold')
    ax_wheel.set_title(f"CIELAB Color Shift Map (Limit: {de_a:.2f} ΔE CMC)", fontsize=9, fontweight='bold', pad=6)
    ax_wheel.axis('off'); ax_bar.axis('off')

    # PREMIUM BYPASS 1: Pre-render the Legend ONCE outside the loop!
    from matplotlib.lines import Line2D
    SENSOR_COLORS = ['#00d2ff', '#ff007c', '#00ff88', '#ffcc00', '#b800ff']
    legend_elements = [Line2D([0], [0], marker='o', color='w', label=f'S{s}', markerfacecolor=SENSOR_COLORS[s], markersize=6, markeredgecolor='black', mew=0.5) for s in range(5) if active_sensors[s]]
    ax_wheel.legend(handles=legend_elements, loc='upper left', fontsize=4.5, markerscale=0.6, handletextpad=0.2, borderpad=0.2, labelspacing=0.2, framealpha=0.6, edgecolor='none')

    # PREMIUM BYPASS 2: Pre-allocate Excel Styles for massive memory savings!
    STYLE_PURPLE_BOLD = Font(bold=True, color="7030A0", size=13)
    STYLE_BLACK_BOLD = Font(bold=True, size=13)
    STYLE_RED_BOLD = Font(bold=True, color="FF0000", size=13)
    STYLE_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i in range(total_points):
        current_dist = dists[i]
        is_last_point = (i == total_points - 1)

        # =====================================================================
        # --- PREMIUM FIX: THE DUAL-STATE ISOLATION ENGINE ---
        # Separates clean fabric from stains by triggering on BOTH start and end!
        # =====================================================================
        
        # =====================================================================
        # --- PREMIUM FIX: THE MICRO-STAIN ISOLATOR ---
        # =====================================================================
        
        # 1. Evaluate true physical state (PREMIUM FIX: Generator Bypass!)
        # Eradicates millions of temporary generator objects from RAM for pure speed.
        currently_failing = False
        for s in range(5):
            if active_sensors[s] and s_data_list[s][i] >= de_a:
                currently_failing = True
                break
                
        trigger_print = False

        # RULE 1: Standard 2-Yard Checkpoint
        if current_dist >= next_interval_target:
            trigger_print = True
        
        # RULE 2: The Chatter-Immune Boundary
        # We ONLY seal the block if the fabric has physically rolled forward.
        # This completely eradicates zero-distance sensor chatter while flawlessly
        # isolating microscopic needle-spikes!
        if i > 0 and currently_failing != block_state_failing:
            if (current_dist - interval_start_dist) > 0.002: 
                trigger_print = True

        if trigger_print or is_last_point:
            
            # =====================================================================
            # --- PREMIUM FIX: THE SMART EXCEL FILTER & INDEPENDENT TRACKERS ---
            # Evaluate if this block is worth printing to the final report.
            # We ONLY print if it is a FAIL stain, a 2-Yard Checkpoint, or the final row!
            # =====================================================================
            is_2_yard_checkpoint = (current_dist >= next_interval_target)
            is_fail_block = any(running_max_de[s] >= de_a for s in range(5))
            should_print = is_fail_block or is_2_yard_checkpoint or is_last_point
            
            if should_print:
                if ws2.max_row > 1: ws2.append([])
                ws2.append([f"{interval_start_dist:.3f} Yards - {current_dist:.3f} Yards"])
                start_row = ws2.max_row
                ws2.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
                title_cell = ws2.cell(row=start_row, column=1)
                title_cell.alignment = center_alignment; title_cell.font = Font(bold=True, color="FFFFFF", size=16); title_cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
                
                sensor_row = start_row + 1
                row_sensors = []
                for s in range(5): row_sensors.extend([f"Sensor #{s}", ""])
                ws2.append(row_sensors)
                for s in range(5):
                    col_start = s * 2 + 1
                    ws2.merge_cells(start_row=sensor_row, start_column=col_start, end_row=sensor_row, end_column=col_start+1)
                    s_cell = ws2.cell(row=sensor_row, column=col_start)
                    s_cell.alignment = center_alignment; s_cell.font = Font(bold=True, color="FFFFFF", size=14); s_cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
                
                dL_row = []; da_row = []; db_row = []; dC_row = []; dH_row = []; dE_row = []; var_row = []; pf_row = []
                dH_letters = []
                for s in range(5):
                    if active_sensors[s]:
                        # --- PREMIUM FIX: Pull the absolute worst snapshot for THIS specific sensor! ---
                        s_target = running_target_idx[s]
                        
                        m_a = master_cache[s]['m_a']; m_b = master_cache[s]['m_b']; m_C = master_cache[s]['m_C']
                        
                        s_dL = det_data[s]['dL'][s_target]
                        s_da = det_data[s]['da'][s_target]
                        s_db = det_data[s]['db'][s_target]
                        worst_interval_dE = s_data_list[s][s_target]
                        
                        # PREMIUM FIX: Replaced expensive CPU division with cached O(1) inverse multiplication!
                        s_ratio = worst_interval_dE * INV_SIGMA 
                        s_var = 100.0 - (100.0 * math.exp(-0.5 * (s_ratio * s_ratio)))
                        samp_a = m_a + s_da; samp_b = m_b + s_db
                        samp_C = math.sqrt((samp_a * samp_a) + (samp_b * samp_b))
                        s_dC = samp_C - m_C
                        s_dH = math.sqrt(max(0.0, ((s_da * s_da) + (s_db * s_db)) - (s_dC * s_dC)))
                        
                        if m_C > 0.5:
                            sign_dH = 1.0 if ((m_a * samp_b) - (m_b * samp_a)) >= 0 else -1.0
                            s_dH_signed = s_dH * sign_dH
                            m_h_deg = master_cache[s]['m_h_deg']
                            if sign_dH > 0:
                                if 0 <= m_h_deg < 90: h_let = "Y"
                                elif 90 <= m_h_deg < 180: h_let = "G"
                                elif 180 <= m_h_deg < 270: h_let = "B"
                                else: h_let = "R"
                            else:
                                if 0 <= m_h_deg <= 90: h_let = "R"
                                elif 90 < m_h_deg <= 180: h_let = "Y"
                                elif 180 < m_h_deg <= 270: h_let = "G"
                                else: h_let = "B"
                        else:
                            s_dH_signed = 0.0; h_let = ""

                        pf_status = "FAIL" if worst_interval_dE >= de_a else "PASS"
                        dL_row.extend(["dL", round(s_dL, 2)]); da_row.extend(["da", round(s_da, 2)]); db_row.extend(["db", round(s_db, 2)])
                        dC_row.extend(["dC", round(s_dC, 2)]); dH_row.extend(["dH", round(s_dH_signed, 2)]); dE_row.extend(["ΔE CMC", round(worst_interval_dE, 2)])
                        var_row.extend(["Variation", round(s_var, 2)]); pf_row.extend(["P/F ΔE CMC", f"{de_a_str} ({pf_status})"])
                        dH_letters.append(h_let)
                    else:
                        dL_row.extend(["dL", "OFF"]); da_row.extend(["da", "OFF"]); db_row.extend(["db", "OFF"]); dC_row.extend(["dC", "OFF"]); dH_row.extend(["dH", "OFF"]); dE_row.extend(["ΔE", "OFF"]); var_row.extend(["Variation", "OFF"]); pf_row.extend(["P/F ΔE", "OFF"])
                        dH_letters.append("")
                
                ws2.append(dL_row); ws2.append(da_row); ws2.append(db_row); ws2.append(dC_row); ws2.append(dH_row); ws2.append(dE_row); ws2.append(var_row); ws2.append(pf_row)
                
                for r in range(start_row, start_row + 10):
                    for c in range(1, 11):
                        cell = ws2.cell(row=r, column=c); cell.border = thin_border; cell.alignment = center_alignment
                        if r >= start_row + 2:
                            if c % 2 != 0: cell.font = STYLE_PURPLE_BOLD 
                            else:
                                val = cell.value; s_idx = (c // 2) - 1
                                if val != "OFF":
                                    cell.font = STYLE_BLACK_BOLD 
                                    if r == start_row + 2: cell.number_format = '0.00" L";-0.00" D";0.00'
                                    elif r == start_row + 3: cell.number_format = '[Red]0.00" R";[Color10]-0.00" G";0.00'
                                    elif r == start_row + 4: cell.number_format = '[Color44]0.00" Y";[Blue]-0.00" B";0.00'
                                    elif r == start_row + 5: cell.number_format = '0.00" B";-0.00" D";0.00'
                                    elif r == start_row + 6:
                                        let = dH_letters[s_idx]
                                        if let == "R": cell.number_format = '[Red]0.00" R";[Red]-0.00" R";0.00'
                                        elif let == "G": cell.number_format = '[Color10]0.00" G";[Color10]-0.00" G";0.00'
                                        elif let == "Y": cell.number_format = '[Color44]0.00" Y";[Color44]-0.00" Y";0.00'
                                        elif let == "B": cell.number_format = '[Blue]0.00" B";[Blue]-0.00" B";0.00'
                                    elif r == start_row + 7:
                                        cell.number_format = '0.00'
                                        if float(val) >= de_a: cell.font = STYLE_RED_BOLD; cell.fill = STYLE_RED_FILL
                                    elif r == start_row + 8: cell.number_format = '0.00"%"'
                                    elif r == start_row + 9:
                                        if "FAIL" in str(val): cell.font = STYLE_RED_BOLD

                # =====================================================================
                # --- THE 300-DPI DOT PLOTTER (PREMIUM O(1) ARRAY BYPASS) ---
                # =====================================================================
                # Draws the worst-case moment independently for each sensor!
                drawn_artists = [] 
                for s in range(5):
                    if active_sensors[s]:
                        s_target = running_target_idx[s]
                        s_dL = det_data[s]['dL'][s_target]
                        s_da = det_data[s]['da'][s_target]
                        s_db = det_data[s]['db'][s_target]
                        color = SENSOR_COLORS[s]
                        
                        p1, = ax_wheel.plot(s_da, s_db, marker='o', color=color, markersize=5, markeredgecolor='black', mew=0.5)
                        p2, = ax_bar.plot(0, s_dL, marker='o', color=color, markersize=5, markeredgecolor='black', mew=0.5)
                        drawn_artists.extend([p1, p2])

                buf = io.BytesIO()
                report_fig.savefig(buf, format='png', bbox_inches='tight', pad_inches=0.1, transparent=False, facecolor='white')
                buf.seek(0)
                
                for artist in drawn_artists:
                    artist.remove()
                    
                img = XLImage(buf); img.width = 786; img.height = 393; img.anchor = f"L{start_row}"
                ws2.add_image(img)
                for r_idx in range(start_row, start_row + 10): ws2.row_dimensions[r_idx].height = 29.5

            # =====================================================================
            # --- TRACKER RESET (ALWAYS RUNS, EVEN IF PRINT IS SKIPPED!) ---
            # =====================================================================
            
            # Advance the 2-yard window safely
            if current_dist >= next_interval_target:
                next_interval_target = (math.floor(current_dist * 0.5) + 1) * 2.0

            # Lock in the new start distance and reset peak tracker for the next block
            interval_start_dist = current_dist 
            running_max_de = [-1.0] * 5
            
            # --- PREMIUM FIX: Lock in the official state of the NEW block! ---
            block_state_failing = currently_failing

        # =====================================================================
        # --- THE POST-PRINT ABSORBER (5-LANE INDEPENDENT) ---
        # =====================================================================
        for s in range(5):
            if active_sensors[s] and s_data_list[s][i] > running_max_de[s]:
                running_max_de[s] = s_data_list[s][i]
                running_target_idx[s] = i

    for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']: ws2.column_dimensions[letter].width = 13

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return base64.b64encode(output.read()).decode('utf-8'), default_name

# ==========================================
# 3. WEBSOCKET BROADCASTER (Live Graph Pipe)
# ==========================================
class ConnectionManager:
    def __init__(self):
        self.active_connections: list[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
        print(f"[SYSTEM] Web Dashboard Connected! Total clients: {len(self.active_connections)}")

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)
        print("[SYSTEM] Web Dashboard Disconnected.")

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            try:
                await connection.send_text(message)
            except:
                pass

manager = ConnectionManager()

@app.websocket("/ws/graph")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            data = await websocket.receive_text()
            packet = json.loads(data)
            
            if packet.get("type") == "control":
                cmd = packet.get("command")
                
                # Instantly update server math configurations based on UI sliders
                global CURRENT_SIGMA, INV_SIGMA, active_sensors
                if "sigma" in packet: 
                    CURRENT_SIGMA = float(packet["sigma"])
                    INV_SIGMA = 1.0 / CURRENT_SIGMA  # <--- Updates the inverse cache instantly!
                if "active_sensors" in packet: active_sensors = packet["active_sensors"]

                # Handle Local Cloud Commands
                if cmd == "CLEAR":
                    history_dist.clear()
                    history_rms.clear()
                    history_worst.clear()
                    for i in range(5):
                        history_sensor[i].clear()
                        history_details[i]['dL'].clear()
                        history_details[i]['da'].clear()
                        history_details[i]['db'].clear()
                    print("\n[VAULT] Cloud Data History Cleared!")
                
                # Route Hardware Commands to the ESP32!
                elif cmd == "C_TRIGGER":
                    mode = packet.get("mode", "MASTER")
                    
                    if mode == "MASTER":
                        mqtt_client.publish(MQTT_TOPIC_PUB, "2")
                    elif mode == "SELECTIVE":
                        # Triggers the first active sensor to start the sequence
                        first_s = next((i for i, a in enumerate(active_sensors) if a), 0)
                        mqtt_client.publish(MQTT_TOPIC_PUB, f"C{first_s}")
                    elif mode == "MANUAL":
                        mqtt_client.publish(MQTT_TOPIC_PUB, "3")
                        
                    print(f"[ROUTER] Triggered Calibration Mode: {mode}")
                else:
                    mqtt_client.publish(MQTT_TOPIC_PUB, f"{cmd}")
                    print(f"[ROUTER] Sent to ESP32: {cmd}")
            
            elif packet.get("type") == "export_excel":
                print("\n[VAULT] Generating massive dual-sheet Excel report...")
                try:
                    b64_data, filename = generate_cloud_excel()
                    await websocket.send_text(json.dumps({
                        "type": "excel_file",
                        "file_data": b64_data,
                        "filename": filename
                    }))
                    print(f"[VAULT] Excel {filename} sent to browser!")
                except Exception as e:
                    print(f"[ERROR] Excel Generation Failed: {e}")
                    
    except WebSocketDisconnect:
        manager.disconnect(websocket)

# ==========================================
# 4. AWS MQTT LISTENER
# ==========================================
def on_connect(client, userdata, flags, rc):
    print("[SUCCESS] Engine Handshake with AWS IoT Core Complete!")
    client.subscribe(MQTT_TOPIC_SUB)
    print(f"[SUCCESS] Listening for JSON data on: {MQTT_TOPIC_SUB}")

def on_message(client, userdata, msg):
    payload = msg.payload.decode('utf-8')
    try:
        data = json.loads(payload)
        
        # --- RESTORED: CLOUD VAULT CATCHER ---
        if data.get("type") == "calib":
            global CURRENT_CALIB_MODE
            CURRENT_CALIB_MODE = data.get("mode", "MASTER")
            master_vals = data["m"]
            
            for i in range(5):
                if i < len(master_vals):
                    vault_data[i]['L'] = master_vals[i][0]
                    vault_data[i]['a'] = master_vals[i][1]
                    vault_data[i]['b'] = master_vals[i][2]
            
            # Clear the math cache since the master standards just changed!
            global _cmc_weight_cache
            _cmc_weight_cache.clear()
            
            print(f"\n[VAULT] 📥 New {CURRENT_CALIB_MODE} Calibration securely saved in Cloud RAM!")
            return # Stop here, don't run the normal distance math

        # --- NORMAL DATA ROUTING ---
        dist = data['dist']
        s_deltas = data['s']

        deltas = []
        active_list = [] # Tracker for strictly active sensors
        
        for i in range(5):
            if active_sensors[i]:
                active_list.append(i)
                dL, da, db = s_deltas[i]
                L_std = vault_data[i]['L']
                a_std = vault_data[i]['a']
                b_std = vault_data[i]['b']
                de_cmc, _, _, _ = calc_cmc(L_std, a_std, b_std, dL, da, db)
                deltas.append(de_cmc)
            else:
                deltas.append(0.0)

        # Statistical Calculations (PREMIUM RMS LOGIC SYNCED WITH GEM_FINAL2)
        if active_list:
            max_delta = max(deltas[idx] for idx in active_list)
            
            rms_list = [deltas[idx] for idx in active_list]
            # The Critical Feature: Reject massive outlier spikes from skewing the Master RMS
            if max_delta > CURRENT_SIGMA and len(rms_list) > 1:
                rms_list.remove(max_delta)
                
            # --- PREMIUM FIX: O(1) Multiplication, Zero Division ---
            inv_len = 1.0 / len(rms_list)
            mean_square = sum((d * d) for d in rms_list) * inv_len
            master_rms = math.sqrt(mean_square)
        else:
            max_delta = 0.0
            master_rms = 0.0

        # --- PREMIUM FIX: O(1) Multiplication using Global Cache ---
        rms_ratio = master_rms * INV_SIGMA
        sqi_rms = 100.0 * math.exp(-0.5 * (rms_ratio * rms_ratio))

        # --- NEW: SAVE TO CLOUD HISTORY VAULT ---
        history_dist.append(dist)
        history_rms.append(master_rms)
        history_worst.append(max_delta)
        for i in range(5):
            history_sensor[i].append(deltas[i])
            if active_sensors[i]:
                history_details[i]['dL'].append(s_deltas[i][0])
                history_details[i]['da'].append(s_deltas[i][1])
                history_details[i]['db'].append(s_deltas[i][2])

        # Build Web Payload
        web_payload = {
            "dist": dist,
            "rms": round(master_rms, 3),
            "worst_de": round(max_delta, 3),
            "sqi": round(sqi_rms, 1),
            "sensors": [round(d, 3) for d in deltas]
        }

        # Blast it to the Web Dashboard!
        asyncio.run(manager.broadcast(json.dumps(web_payload)))
        print(f"[DATA] Dist: {dist:.3f} Yds | Master RMS: {master_rms:.2f} | SQI: {sqi_rms:.1f}%")

    except Exception as e:
        print(f"[ERROR] Bad JSON Payload: {e}")

mqtt_client = mqtt.Client(client_id="")
mqtt_client.tls_set(ca_certs="certs/root-CA.pem",
                    certfile="certs/certificate.pem.crt",
                    keyfile="certs/private.pem.key",
                    tls_version=ssl.PROTOCOL_TLSv1_2)
mqtt_client.on_connect = on_connect
mqtt_client.on_message = on_message

@app.on_event("startup")
async def startup_event():
    if AWS_ENDPOINT != "YOUR_AWS_ENDPOINT_HERE":
        mqtt_client.connect(AWS_ENDPOINT, 8883, 60)
        mqtt_client.loop_start()
    else:
        print("\n[FATAL ERROR] Check AWS ENDPOINT!\n")


# ==========================================
# THE GLOBAL STATIC MOUNTER
# ==========================================
import os
from fastapi.staticfiles import StaticFiles

app.mount("/", StaticFiles(directory=os.path.dirname(os.path.abspath(__file__)), html=True), name='static')
