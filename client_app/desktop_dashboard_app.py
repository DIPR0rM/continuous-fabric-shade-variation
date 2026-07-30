import customtkinter as ctk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import serial
import time
import sys
import threading
import collections
import math
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# --- NEW: Moved from inside the Excel loop to the top for maximum speed ---
import io
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.patches import Circle, Rectangle, Ellipse, Polygon
from openpyxl.drawing.image import Image as XLImage
import matplotlib.gridspec as gridspec

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

COM_PORT = "Your_COM"  
BAUD_RATE = 500000
VISUAL_BUFFER_LIMIT = 2000  

SENSOR_COLORS = ['#00d2ff', '#ff007c', '#00ff88', '#ffcc00', '#b800ff']

# --- GLOBAL MATH CACHE: Pre-calculate logarithmic constants once at boot ---
SIGMA_FACTOR_A = math.sqrt(-2 * math.log(0.94)) # 6% Variation Line
SIGMA_FACTOR_B = math.sqrt(-2 * math.log(0.89)) # 11% Variation Line

class TMDMDashboard(ctk.CTk):
    def __init__(self, ser_instance, active_sensors, calib_mode):
        super().__init__()
        self.title("TMDM Shade Inspection - Premium Graphics Engine")
        self.geometry("1650x850") 

        self.ser = ser_instance
        self.is_running = True
        self.is_paused = True 
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.data_lock = threading.Lock()

        # --- SELECTIVE CALIBRATION ENGINE VARIABLES ---
        self.selective_active_list = [idx for idx, state in enumerate(active_sensors) if state]
        self.selective_current_idx = 0
        self.allow_selective_capture = False
        # ----------------------------------------------

        self.all_distances = []
        self.all_rms = [] 
        self.all_worst_de = [] 
        self.all_sensor_data = {0: [], 1: [], 2: [], 3: [], 4: []}
        self.all_detailed_data = {i: {'dL': [], 'da': [], 'db': []} for i in range(5)}
        
        self.view_distances = collections.deque(maxlen=VISUAL_BUFFER_LIMIT)
        self.rms_snapshot = collections.deque(maxlen=VISUAL_BUFFER_LIMIT) 
        self.view_sensor_data = {i: collections.deque(maxlen=VISUAL_BUFFER_LIMIT) for i in range(5)}
        
        self.splice_offset_yards = 0.0
        self.last_known_distance = 0.0
        
        self.active_sensors_state = active_sensors
        self.active_sensor_indices = [idx for idx, state in enumerate(active_sensors) if state] # HIGH-SPEED RAM CACHE
        self.current_sigma_val = 1.50
        
        self.calib_mode = calib_mode
        self.vault_data = {i: {'L': 0.0, 'a': 0.0, 'b': 0.0} for i in range(5)}
        
        self.current_max_de = 0.0
        self.current_worst_sensor = 0
        
        self.record_min_var = 100.0
        self.record_max_var = 0.0
        self.record_max_de = 0.0
        self.record_max_sensor = 0
        self.record_max_dist = 0.0
        
        # --- NEW: Duration Trackers ---
        self.total_running_seconds = 0.0
        self.total_paused_seconds = 0.0
        self.last_timer_tick = time.time()
        
        self.last_physical_distance = 0.0
        self.last_movement_time = time.time()
        
        self.last_sigma = -1.0 
        self.last_stats_text = ""
        self.last_live_text = ""
        self.p99_variation_text = "" 
        
        self.calibration_flag = False
        self.calibration_time_str = ""

        # --- THE THREAD-SAFE MEMORY VAULT ---
        self.pending_qtx_prompt = ""
        self.pending_saved_message = ""
        self.pending_hard_reset = False
        self.pending_soft_reset = False

        self.setup_ui()
        self.hardware_thread = threading.Thread(target=self.serial_worker_thread, daemon=True)
        self.hardware_thread.start()
        self.after(100, self.update_graphics_loop)

    def on_closing(self):
        print("\n[SYSTEM] Closing Dashboard... Restarting Bootloader.")
        self.is_running = False  
        
        # FIX: Unlock the COM port so the terminal can use it again!
        if self.ser and self.ser.is_open:
            try:
                self.ser.close()
            except:
                pass
                
        self.quit()
        self.destroy()

    def setup_ui(self):
        self.title_label = ctk.CTkLabel(self, text="Real-Time Fabric Shade Analysis", font=ctk.CTkFont(size=26, weight="bold"))
        self.title_label.pack(pady=5)

        self.fig, (self.ax1, self.ax2) = plt.subplots(1, 2, figsize=(16, 6.2), facecolor='#2b2b2b', gridspec_kw={'width_ratios': [1.8, 1]})
        self.fig.tight_layout(pad=4.0)

        self.ax1.set_facecolor('#2b2b2b')
        self.ax1.set_title("Continuous Hardware Scan (Active Sensors Only)", color='white', fontsize=12)
        self.ax1.set_xlabel("Distance (Yards)", color='white')
        self.ax1.set_ylabel("ΔE CMC", color='white')
        self.ax1.tick_params(colors='white')
        
        self.lines = []
        for i in range(5):
            line, = self.ax1.plot([], [], color=SENSOR_COLORS[i], linewidth=2.5, label=f'Sensor {i}')
            self.lines.append(line)
            
        self.rms_line, = self.ax1.plot([], [], color='white', linewidth=3, linestyle='--', label='Master RMS')
        self.sigma_line = self.ax1.axhline(y=1.5, color='white', linestyle='--', linewidth=1.5, alpha=0.6, label='Tolerance (σ)')
        self.ax1.legend(facecolor='#2b2b2b', edgecolor='white', labelcolor='white', loc='upper right')

        self.live_data_box = self.ax1.text(0.02, 0.96, "Waiting for data...", transform=self.ax1.transAxes, fontsize=11, color='white', verticalalignment='top', horizontalalignment='left', bbox=dict(boxstyle='round,pad=0.5', facecolor='#2b2b2b', edgecolor='white', alpha=0.8))
        self.calib_box = self.ax1.text(0.50, 0.96, f"Calibration Mode: {self.calib_mode}", transform=self.ax1.transAxes, fontsize=11, color='#00ff88', verticalalignment='top', horizontalalignment='center', bbox=dict(boxstyle='round,pad=0.5', facecolor='#2b2b2b', edgecolor='white', alpha=0.8))

        self.ax2.set_facecolor('#2b2b2b')
        self.ax2.set_title("Statistical Shade Quality Index (SQI)", color='white', fontsize=12)
        self.ax2.set_xlabel("Color Difference (ΔE CMC)", color='white')
        self.ax2.set_ylabel("Quality Score (%)", color='white')
        self.ax2.tick_params(colors='white')
        self.ax2.set_xlim(0, 5)
        self.ax2.set_ylim(0, 105)
        self.line_curve, = self.ax2.plot([], [], color='#ff007c', linewidth=2.5)
        
        self.dot_current, = self.ax2.plot([], [], 'o', color='white', markersize=9, label='Master RMS')
        self.dot_worst, = self.ax2.plot([], [], 'o', color='#ff3333', markersize=7, label='Worst Sensor')
        self.ax2.legend(facecolor='#2b2b2b', edgecolor='white', labelcolor='white', loc='upper right')
        
        # --- FIX: Move Rainbow Line to the Left Graph (ax1) and make it Horizontal ---
        gradient_h = np.linspace(0, 1, 256).reshape(1, -1)
        self.rainbow_line = self.ax1.imshow(gradient_h, cmap='rainbow', extent=[0, 10, 0, 0], aspect='auto', origin='lower', alpha=0.8, zorder=2)
        
        self.stats_box = self.ax2.text(0.95, 0.82, "Min Variation: --%\nMax Variation: --%\nWorst ΔE CMC: --\nMax ΔE CMC: --\n(System Ready. Press Get Calibration.)", transform=self.ax2.transAxes, fontsize=11, color='white', verticalalignment='top', horizontalalignment='right', bbox=dict(boxstyle='round,pad=0.5', facecolor='#2b2b2b', edgecolor='white', alpha=0.8))

        math_formula = r"$SQI = 100 \times e^{-\frac{1}{2}\left(\frac{\Delta E_{cmc}}{\sigma}\right)^2}$"
        self.formula_box = self.ax2.text(0.05, 0.05, math_formula, transform=self.ax2.transAxes, fontsize=14, color='#00d2ff', verticalalignment='bottom', horizontalalignment='left', bbox=dict(boxstyle='round,pad=0.5', facecolor='#2b2b2b', edgecolor='white', alpha=0.8))

        # --- NEW UI: Live Hardware Timer Box (Moved to Left) ---
        self.timer_box = self.ax2.text(0.05, 0.18, "Running: 0m 0s\nPaused: 0m 0s", transform=self.ax2.transAxes, fontsize=11, color='white', verticalalignment='bottom', horizontalalignment='left', bbox=dict(boxstyle='round,pad=0.5', facecolor='#2b2b2b', edgecolor='white', alpha=0.8))
        self.last_timer_text = ""

        self.canvas = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().pack(fill=ctk.BOTH, expand=True, padx=20, pady=5)

        self.control_frame = ctk.CTkFrame(self)
        self.control_frame.pack(fill=ctk.X, padx=20, pady=5)

        self.control_row1 = ctk.CTkFrame(self.control_frame, fg_color="transparent")
        self.control_row1.pack(fill=ctk.X, padx=5, pady=5)

        self.btn_start = ctk.CTkButton(self.control_row1, text="Get New Calibration Data", command=self.trigger_calibration, fg_color="#00cc66", hover_color="#00994c", width=160)
        self.btn_start.pack(side=ctk.LEFT, padx=5)

        self.btn_pause = ctk.CTkButton(self.control_row1, text="Pause (0)", command=self.pause_machine, fg_color="#cc9900", hover_color="#b38600", width=80)
        self.btn_pause.pack(side=ctk.LEFT, padx=5)

        self.btn_resume = ctk.CTkButton(self.control_row1, text="Resume (1)", command=self.resume_machine, fg_color="#0088cc", hover_color="#006699", width=80)
        self.btn_resume.pack(side=ctk.LEFT, padx=5)

        self.btn_clear = ctk.CTkButton(self.control_row1, text="Clear Data", command=self.hard_clear_data, fg_color="#e60000", hover_color="#b30000", width=80)
        self.btn_clear.pack(side=ctk.LEFT, padx=5)

        self.btn_export = ctk.CTkButton(self.control_row1, text="Export Excel", command=self.export_to_excel, fg_color="#b800ff", hover_color="#8a00c2", width=100)
        self.btn_export.pack(side=ctk.LEFT, padx=15)

        self.sensor_frame = ctk.CTkFrame(self.control_row1, fg_color="transparent")
        self.sensor_frame.pack(side=ctk.LEFT, padx=15)
        
        self.sensor_checkbox_vars = []
        for i in range(5):
            var = ctk.BooleanVar(value=self.active_sensors_state[i])
            self.sensor_checkbox_vars.append(var)
            cb = ctk.CTkCheckBox(self.sensor_frame, text=f"S{i}", variable=var, command=self.update_sensor_state, width=45, checkbox_height=18, checkbox_width=18)
            cb.pack(side=ctk.LEFT, padx=2)
            
            # PERFECT SYNC: Uses the fixed bootloader state!
            if self.active_sensors_state[i]: cb.select()
            else: cb.deselect()

        self.control_row2 = ctk.CTkFrame(self.control_frame, fg_color="transparent")
        self.control_row2.pack(fill=ctk.X, padx=5, pady=5)

        # --- FIX 1: Pack the Readout FIRST so it claims its space and never gets clipped ---
        self.sqi_readout = ctk.CTkLabel(self.control_row2, text="RMS Var: --% | Worst: --%", font=ctk.CTkFont(size=14, weight="bold"), text_color="#00d2ff")
        self.sqi_readout.pack(side=ctk.RIGHT, padx=10)

        # FIXED UI BUG: Static placeholder completely prevents the overlap/minimize glitch
        self.terminal_entry = ctk.CTkEntry(self.control_row2, placeholder_text="Enter Command or Value...", width=250)
        self.terminal_entry.pack(side=ctk.LEFT, padx=5)
        self.terminal_entry.bind("<Return>", self.send_terminal_command)

        self.btn_send = ctk.CTkButton(self.control_row2, text="Send", command=self.send_terminal_command, fg_color="#444444", hover_color="#555555", width=50)
        self.btn_send.pack(side=ctk.LEFT, padx=5)

        # --- FIX 2: Reduce the massive dead-space width from 450 to 280 ---
        self.qtx_prompt_label = ctk.CTkLabel(self.control_row2, text=" [ Standby ] ", font=ctk.CTkFont(size=14, weight="bold"), text_color="#ffcc00", width=280, anchor="w")
        self.qtx_prompt_label.pack(side=ctk.LEFT, padx=10)

        # --- NEW UI: Passed ΔE Box ---
        self.passed_de_label = ctk.CTkLabel(self.control_row2, text="Passed ΔE CMC:", font=ctk.CTkFont(size=13, weight="bold"))
        self.passed_de_label.pack(side=ctk.LEFT, padx=(15, 2))

        self.passed_de_entry = ctk.CTkEntry(self.control_row2, width=55)
        self.passed_de_entry.pack(side=ctk.LEFT, padx=5)
        self.passed_de_entry.bind("<Return>", self.update_from_passed_de)
        self.passed_de_entry.bind("<FocusOut>", self.update_from_passed_de)

        self.slider_label = ctk.CTkLabel(self.control_row2, text="Tolerance (σ): 1.50", font=ctk.CTkFont(size=13))
        self.slider_label.pack(side=ctk.LEFT, padx=10)
        
        self.sigma_slider = ctk.CTkSlider(self.control_row2, from_=0.5, to=10.0, number_of_steps=95, command=self.slider_event, width=130)
        self.sigma_slider.set(1.5)
        self.sigma_slider.pack(side=ctk.LEFT, padx=5)
        
        # Initialize the Passed DE box correctly for 1.50 Sigma
        initial_passed_de = 1.5 * SIGMA_FACTOR_A
        self.passed_de_entry.insert(0, f"{initial_passed_de:.2f}")

    # ========================================================================
    # 4. MASTER CONTROL LOGIC (GUI State Memory)
    # ========================================================================
    def update_sensor_state(self):
        self.active_sensors_state = [var.get() for var in self.sensor_checkbox_vars]
        self.active_sensor_indices = [idx for idx, state in enumerate(self.active_sensors_state) if state] # UPDATE RAM CACHE
        
        # Instantly force the Live Data box to reflect 'Off' changes
        with self.data_lock:
            sensor_snapshot = {i: list(self.view_sensor_data[i]) for i in range(5)}
            dist_snapshot = list(self.view_distances)
            rms_snap = list(self.rms_snapshot)
        
        if len(dist_snapshot) > 0:
            current_rms_val = rms_snap[-1] if len(rms_snap) > 0 else 0.0
            latest_vals_str = []
            for i in range(5):
                if self.active_sensors_state[i]:
                    val = sensor_snapshot[i][-1] if len(sensor_snapshot[i]) > 0 else 0.0
                    latest_vals_str.append(f"Sensor {i}: {val:.2f}")
                else:
                    latest_vals_str.append(f"Sensor {i}: Off")
                    
            new_live_text = f"Distance: {dist_snapshot[-1]:.3f} Yards\n-----------------\nCurrent ΔE CMC:\n" + "\n".join(latest_vals_str) + f"\nMaster RMS: {current_rms_val:.2f}"
            self.live_data_box.set_text(new_live_text)
            self.last_live_text = new_live_text
            self.fig.canvas.draw_idle()

    # ========================================================================
    # THE PROFESSIONAL CMC (2:1) TEXTILE ALGORITHM (LEVEL-2 CACHE OPTIMIZED)
    # ========================================================================
    def calc_cmc(self, L_std, a_std, b_std, dL, da, db, l=2.0, c=1.0):
        # 1. SAMPLE MAGNITUDES (Always unique, must be calculated live)
        C_std = math.sqrt((a_std * a_std) + (b_std * b_std))
        a_samp = a_std + da
        b_samp = b_std + db
        C_samp = math.sqrt((a_samp * a_samp) + (b_samp * b_samp))
        
        dC = C_samp - C_std
        dH_sq = max(0.0, ((da * da) + (db * db)) - (dC * dC)) 

        # 2. THE MASTER CACHE (Bypasses trigonometry and division if Master is unchanged)
        cache_key = (L_std, a_std, b_std)
        if not hasattr(self, '_cmc_weight_cache'):
            self._cmc_weight_cache = {}
            
        if cache_key in self._cmc_weight_cache:
            # INSTANT RAM PULL: Skip all heavy math
            inv_L_weight, inv_C_weight, inv_H_sq_weight, S_C, S_H, S_L = self._cmc_weight_cache[cache_key]
        else:
            # HEAVY CALCULATION: Only runs once per Master Calibration
            S_L = 0.511 if L_std < 16.0 else (0.040975 * L_std) / (1.0 + 0.01765 * L_std)
            S_C = ((0.0638 * C_std) / (1.0 + 0.0131 * C_std)) + 0.638
            
            if C_std < 0.0001: 
                S_H = S_C 
            else:
                h_std = math.degrees(math.atan2(b_std, a_std))
                if h_std < 0: h_std += 360.0
                
                C_std_sq = C_std * C_std
                C_std_4 = C_std_sq * C_std_sq
                F = math.sqrt(C_std_4 / (C_std_4 + 1900.0))
                
                if 164.0 <= h_std <= 345.0:
                    T = 0.56 + abs(0.2 * math.cos(math.radians(h_std + 168.0)))
                else:
                    T = 0.36 + abs(0.4 * math.cos(math.radians(h_std + 35.0)))
                    
                S_H = S_C * (F * T + 1.0 - F)

            inv_L_weight = 1.0 / (l * S_L)
            inv_C_weight = 1.0 / (c * S_C)
            inv_H_sq_weight = 1.0 / (S_H * S_H)
            
            # Save to RAM for next time
            self._cmc_weight_cache[cache_key] = (inv_L_weight, inv_C_weight, inv_H_sq_weight, S_C, S_H, S_L)

        # 3. HIGH-SPEED ASSEMBLY
        term_L = dL * inv_L_weight
        term_C = dC * inv_C_weight
        term_H_sq = dH_sq * inv_H_sq_weight
        
        de_cmc = math.sqrt((term_L * term_L) + (term_C * term_C) + term_H_sq)
        return de_cmc, S_C, S_H, S_L

    def update_from_passed_de(self, event=None):
        try:
            val = float(self.passed_de_entry.get().strip())
            if val <= 0: return
            
            # MATH: Reverse engineer Sigma from the Passed ΔE value (6% variation curve)
            new_sigma = val / SIGMA_FACTOR_A
            new_sigma = max(0.5, min(10.0, new_sigma)) # Keep it within slider bounds
            
            self.sigma_slider.set(new_sigma)
            self.current_sigma_val = new_sigma
            self.slider_label.configure(text=f"Tolerance (σ): {new_sigma:.2f}")
            if hasattr(self, 'fig'): self.fig.canvas.draw_idle()
        except ValueError:
            pass

    def slider_event(self, value):
        self.current_sigma_val = value
        self.slider_label.configure(text=f"Tolerance (σ): {value:.2f}")
        
        # MATH: Calculate Passed ΔE when user physically moves the slider
        passed_de = value * SIGMA_FACTOR_A
        self.passed_de_entry.delete(0, 'end')
        self.passed_de_entry.insert(0, f"{passed_de:.2f}")

    def send_terminal_command(self, event=None):
        cmd = self.terminal_entry.get().strip()
        if cmd:
            self.send_command(cmd)
            self.terminal_entry.delete(0, 'end')

    def send_command(self, cmd):
        if self.ser:
            if not cmd.endswith('\n'):
                cmd += '\n'
            self.ser.write(cmd.encode('utf-8'))

    def trigger_calibration(self):
        if getattr(self, 'calib_mode', '') == 'MASTER_SELECTIVE':
            if self.selective_current_idx == 0:
                self.selective_active_list = [idx for idx, state in enumerate(self.active_sensors_state) if state]
                
                # --- NEW: Calibration ALWAYS resets distance to 0.0! ---
                self.clear_data(reset_distance=True) 
                self.send_command('Z') 
                
            if self.selective_current_idx >= len(self.selective_active_list):
                self.qtx_prompt_label.configure(text=" ✅ All sensors calibrated. Click 'Clear Data' to restart. ")
                return
            self.allow_selective_capture = True
            
            target_s = self.selective_active_list[self.selective_current_idx]
            
            protected_sensors = self.selective_active_list[:self.selective_current_idx]
            if protected_sensors:
                prot_str = ", ".join([f"S{p}" for p in protected_sensors])
                self.qtx_prompt_label.configure(text=f" ⚙️ Scanning S{target_s}... (Vault Protecting: {prot_str}) ")
            else:
                self.qtx_prompt_label.configure(text=f" ⚙️ Scanning Target Sensor {target_s}... ")
            
            self.send_command(f'C{target_s}')
        else:
            # --- NEW: Calibration ALWAYS resets distance to 0.0! ---
            self.clear_data(reset_distance=True) 
            self.send_command('Z') 
            
            self.qtx_prompt_label.configure(text=" ⚙️ Syncing Hardware Array... ")
            threading.Thread(target=self._invisible_hardware_sync_macro, daemon=True).start()

    def _invisible_hardware_sync_macro(self):
        if not self.ser or not self.ser.is_open:
            return

        # 1. INSTANT SYNC: Send the "U" string to update LEDs without a reboot!
        sync_str = "U" + "".join(['1' if state else '0' for state in self.active_sensors_state])
        self.send_command(sync_str)
        time.sleep(0.2)
        
        # 2. Trigger Target Calibration (Hardware distance remains untouched!)
        self.after(0, lambda: self.qtx_prompt_label.configure(text=""))
        
        if self.calib_mode.startswith('MASTER'):
            self.after(0, lambda: self.calib_box.set_text("Calibrating Master..."))
            self.send_command('2')
        else:
            self.after(0, lambda: self.calib_box.set_text("Awaiting Manual QTX..."))
            self.send_command('3')
        
        self.after(0, lambda: self.fig.canvas.draw_idle())
    

    # GUI Thread-Safe Commands for Serial Background Calls
    def _safe_update_qtx_prompt(self, text):
        self.qtx_prompt_label.configure(text=f" 🎯 {text} ")
        self.calib_box.set_text(text) 
        self.terminal_entry.delete(0, 'end')
        self.terminal_entry.focus_set()
        self.fig.canvas.draw_idle()

    def _safe_show_saved(self, text):
        self.qtx_prompt_label.configure(text=f" ✅ {text} ")

    def _safe_trigger_hard_reset(self):
        print("\n[SYSTEM] ESP32 Hard Reset Wiped! Returning to Bootloader...")
        self.on_closing()

    def _safe_trigger_soft_reset(self):
        self.clear_data()
        self.is_paused = True 
        self.calib_box.set_text("System Cleared. Click 'Get New Calibration'.")
        self.qtx_prompt_label.configure(text=" [ Standby ] ")
        self.terminal_entry.delete(0, 'end')
        self.fig.canvas.draw_idle()

    def pause_machine(self):
        self.send_command('0')
        self.is_paused = True
        
        if len(self.all_rms) > 0:
            total_scans = len(self.all_rms)
            
            # --- DYNAMIC GRADING MATH ---
            sigma_factor_a = SIGMA_FACTOR_A # 6% Variation Line
            sigma_factor_b = SIGMA_FACTOR_B # 11% Variation Line
            de_a = self.current_sigma_val * sigma_factor_a
            de_b = self.current_sigma_val * sigma_factor_b
            
            grade_a_count = sum(1 for rms in self.all_rms if rms <= de_a)
            grade_b_count = sum(1 for rms in self.all_rms if de_a < rms <= de_b)
            grade_c_count = sum(1 for rms in self.all_rms if rms > de_b)
            
            yield_a = (grade_a_count / total_scans) * 100
            yield_b = (grade_b_count / total_scans) * 100
            yield_c = (grade_c_count / total_scans) * 100
            
            report_text = (
                f"Current Grading:\n"
                f"Grade A (<={de_a:.2f}): {yield_a:.1f}%\n"
                f"Grade B (<={de_b:.2f}): {yield_b:.1f}%\n"
                f"Grade C (>{de_b:.2f}) : {yield_c:.1f}%"
            )
            self.stats_box.set_text(report_text)
            
        if len(self.all_worst_de) > 0:
            current_sigma = self.current_sigma_val
            sorted_worst_de = sorted(self.all_worst_de)
            keep_count = int(len(sorted_worst_de) * 0.99)
            if keep_count == 0 and len(sorted_worst_de) > 0: keep_count = 1
            bottom_99_array = sorted_worst_de[:keep_count]
            
            if len(bottom_99_array) > 0:
                true_average_de = sum(bottom_99_array) / len(bottom_99_array)
            else:
                true_average_de = 0.0
            
            # OPTIMIZATION: Native math.exp and direct multiplication
            ratio_99 = true_average_de / current_sigma
            p99_sqi = 100.0 * math.exp(-0.5 * (ratio_99 * ratio_99))
            p99_var = 100.0 - p99_sqi
            
            p100_avg_de = sum(self.all_worst_de) / len(self.all_worst_de)
            ratio_100 = p100_avg_de / current_sigma
            p100_sqi = 100.0 * math.exp(-0.5 * (ratio_100 * ratio_100))
            p100_var = 100.0 - p100_sqi
            
            self.p99_variation_text = f" | P99: {p99_var:.1f}% | P100: {p100_var:.1f}%"
            
        self.fig.canvas.draw_idle()

    def resume_machine(self):
        self.send_command('1')
        self.is_paused = False
        self.p99_variation_text = "" 
        current_text = self.stats_box.get_text()
        if "Current Grading:" in current_text:
            new_text = current_text.replace("Current Grading:", "Previous Grading:")
            self.stats_box.set_text(new_text)
            
        # --- THE MISSING UI FIX ---
        self.qtx_prompt_label.configure(text=" [ Standby ] ")
        # --------------------------
        
        self.fig.canvas.draw_idle()

    def clear_data(self, reset_distance=False):
        with self.data_lock:
            self.all_distances.clear()
            self.view_distances.clear()
            self.all_rms.clear() 
            self.all_worst_de.clear() 
            self.rms_snapshot.clear() 
            self.rms_line.set_data([], []) 
            
            for i in range(5):
                self.all_sensor_data[i].clear()
                self.view_sensor_data[i].clear()
                for key in ['dL', 'da', 'db']:
                    self.all_detailed_data[i][key].clear()
                
            self.current_max_de = 0.0
            self.current_worst_sensor = 0
            self.record_min_var = 100.0
            self.record_max_var = 0.0
            self.record_max_de = 0.0
            self.record_max_sensor = 0
            self.record_max_dist = 0.0
            
            self.total_running_seconds = 0.0
            self.total_paused_seconds = 0.0
            self.last_timer_tick = time.time()
            self.p99_variation_text = "" 
            
            # --- NEW: Only wipe the distance if strictly commanded! ---
            if reset_distance:
                self.last_physical_distance = 0.0
                self.last_movement_time = time.time()
                self.splice_offset_yards = 0.0
                self.last_known_distance = 0.0
            
            for i in range(5):
                self.lines[i].set_data([], [])
            self.ax1.set_xlim(0, 10) 
            self.ax1.set_ylim(0, 5)
            self.stats_box.set_text("Min Variation: --%\nMax Variation: --%\nWorst ΔE: --\nMax ΔE: --\n(Paused - Awaiting Resume)")
            self.live_data_box.set_text("Waiting for data...")
            self.qtx_prompt_label.configure(text=" [ Standby ] ")
            self.fig.canvas.draw_idle()

    def hard_clear_data(self):
        self.selective_current_idx = 0
        self.allow_selective_capture = False
        self.selective_active_list = [idx for idx, state in enumerate(self.active_sensors_state) if state]
        
        # --- NEW: Clearing Data NEVER resets the physical distance! ---
        self.clear_data(reset_distance=False)

    # ========================================================================
    # 5. EXCEL GENERATOR 
    # ========================================================================
    def export_to_excel(self):
        current_sigma = self.current_sigma_val
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        default_name = f"TMDM_Detailed_Report_{timestamp}.xlsx"
        
        # POPUP FIX: Force Windows to open a Save Dialog box
        from tkinter import filedialog
        filename = filedialog.asksaveasfilename(
            parent=self, # <--- THIS FORCES THE POPUP TO THE FRONT!
            defaultextension=".xlsx",
            initialfile=default_name,
            title="Save Excel Report",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        
        if not filename:
            return

        with self.data_lock:
            total_points = len(self.all_distances)
            if total_points == 0:
                self.qtx_prompt_label.configure(text=" ⚠️ No data to export! ")
                self.after(3000, lambda: self.qtx_prompt_label.configure(text=" [ Standby ] "))
                return
            dists = list(self.all_distances)
            s_data_list = [list(self.all_sensor_data[i]) for i in range(5)]
            rms_des = list(self.all_rms)
            worst_des = list(self.all_worst_de)
            det_data = {i: {k: list(self.all_detailed_data[i][k]) for k in ['dL', 'da', 'db']} for i in range(5)}

        # --- DYNAMIC GRADING MATH ---
        sigma_factor_a = SIGMA_FACTOR_A
        sigma_factor_b = SIGMA_FACTOR_B
        de_a = current_sigma * sigma_factor_a
        de_b = current_sigma * sigma_factor_b

        grade_a_count = sum(1 for rms in rms_des if rms <= de_a)
        grade_b_count = sum(1 for rms in rms_des if de_a < rms <= de_b)
        grade_c_count = sum(1 for rms in rms_des if rms > de_b)
        
        yield_a = (grade_a_count / total_points) * 100 if total_points > 0 else 0.0
        yield_b = (grade_b_count / total_points) * 100 if total_points > 0 else 0.0
        yield_c = (grade_c_count / total_points) * 100 if total_points > 0 else 0.0

        if len(worst_des) > 0:
            max_worst_de = max(worst_des)
            worst_index = worst_des.index(max_worst_de)
            max_worst_dist = dists[worst_index]
            
            # --- NEW: Find the exact sensor that caused this worst spike ---
            max_worst_sensor_id = "Unknown"
            for s in range(5):
                if self.active_sensors_state[s] and s_data_list[s][worst_index] == max_worst_de:
                    max_worst_sensor_id = f"#{s}"
                    break
        else:
            max_worst_de = 0.0
            max_worst_dist = 0.0
            max_worst_sensor_id = "N/A"

        sensor_stats = {}
        for i in range(5):
            if self.active_sensors_state[i]:
                s_data = s_data_list[i]
                s_count_a = sum(1 for val in s_data if val <= de_a)
                s_count_b = sum(1 for val in s_data if de_a < val <= de_b)
                s_count_c = sum(1 for val in s_data if val > de_b)
                s_max = max(s_data) if len(s_data) > 0 else 0.0
                s_max_dist = dists[s_data.index(s_max)] if len(s_data) > 0 else 0.0
                
                sensor_stats[i] = {
                    'A': round((s_count_a / total_points) * 100, 1) if total_points > 0 else 0.0,
                    'B': round((s_count_b / total_points) * 100, 1) if total_points > 0 else 0.0,
                    'C': round((s_count_c / total_points) * 100, 1) if total_points > 0 else 0.0,
                    'Max': round(s_max, 2),
                    'MaxDist': round(s_max_dist, 3)
                }
            else:
                sensor_stats[i] = {'A': "N/A", 'B': "N/A", 'C': "N/A", 'Max': "N/A", 'MaxDist': "N/A"}

        if len(worst_des) > 0:
            sorted_worst = sorted(worst_des)
            keep_count = int(len(sorted_worst) * 0.99)
            if keep_count == 0 and len(sorted_worst) > 0: keep_count = 1
            bottom_99_array = sorted_worst[:keep_count]
            if len(bottom_99_array) > 0:
                true_avg_de = sum(bottom_99_array) / len(bottom_99_array)
                # OPTIMIZATION: Removing np.exp and **2 from the Excel Math
                ratio_99 = true_avg_de / current_sigma
                p99_sqi = 100.0 * math.exp(-0.5 * (ratio_99 * ratio_99))
                p99_var = 100.0 - p99_sqi
            else:
                true_avg_de = 0.0
                p99_var = 0.0
                
            p100_avg_de = sum(worst_des) / len(worst_des)
            ratio_100 = p100_avg_de / current_sigma
            p100_sqi = 100.0 * math.exp(-0.5 * (ratio_100 * ratio_100))
            p100_var = 100.0 - p100_sqi
            
            # --- NEW STAT: Outlier Impact Ratio ---
            if true_avg_de > 0:
                outlier_impact = ((p100_avg_de - true_avg_de) / true_avg_de) * 100.0
            else:
                outlier_impact = 0.0
        else:
            p99_var = 0.0
            p100_var = 0.0
            outlier_impact = 0.0

        # --- DYNAMIC TEXT GENERATOR FOR OIR ---
        if outlier_impact <= 2.0:
            oir_warning = "Excellent: Roll is highly consistent. No severe isolated stains detected."
        elif outlier_impact <= 5.0:
            oir_warning = "Notice: Moderate variance. A few noticeable but non-critical spikes exist."
        else:
            oir_warning = "Warning! The base fabric is fine, but there is a massive, severe stain somewhere on this roll that is ruining the average!"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shade Inspection Data"
        fill_A = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") 
        fill_B = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") 
        fill_C = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 
        fill_Grey = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") 

        ws.append(["TMDM Continuous Shade Inspection Final Report"])
        ws["A1"].font = Font(bold=True, size=14)
        
        # --- FIX 1: Merge the title cell across the first 5 columns ---
        ws.merge_cells('A1:E1') 
        
        ws.append(["Export Date/Time:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        # --- FIX 2: Round the Sigma value so it isn't 10 decimal places long ---
        ws.append(["Strictness Tolerance (Sigma):", round(current_sigma, 2)])
        
        ws.append(["Passed Target (ΔE):", round(de_a, 2)])
        ws.append(["Total Fabric Scanned:", f"{dists[-1] if len(dists)>0 else 0.0:.3f} Yards"])
        
        # --- NEW: Format and Append Durations ---
        def fmt_time(secs):
            m, s = divmod(int(secs), 60)
            h, m = divmod(m, 60)
            return f"{h}h {m}m {s}s" if h > 0 else f"{m}m {s}s"
            
        ws.append(["Running Duration:", fmt_time(self.total_running_seconds)])
        ws.append(["Paused Duration:", fmt_time(self.total_paused_seconds)])
        ws.append([])
        
        ws.append(["--- OVERALL ROLL QUALITY (MASTER RMS) ---"])
        r1 = ws.max_row
        ws[f"A{r1}"].font = Font(bold=True, color="00529B") 
        ws.merge_cells(f"A{r1}:E{r1}")
        
        # --- FIX: Send raw numbers, but use Excel's native cell formatting to display the % ---
        ws.append(["Final Yield (Grade A) %:", round(yield_a, 1)])
        ws[f"B{ws.max_row}"].number_format = '0.0"%"'
        
        ws.append(["Final Yield (Grade B) %:", round(yield_b, 1)])
        ws[f"B{ws.max_row}"].number_format = '0.0"%"'
        
        ws.append(["Final Yield (Grade C) %:", round(yield_c, 1)])
        ws[f"B{ws.max_row}"].number_format = '0.0"%"'
        
        ws.append(["Average Worst Variation (P99) %:", round(p99_var, 1)])
        ws[f"B{ws.max_row}"].number_format = '0.0"%"'
        
        ws.append(["Absolute Worst Variation (P100) %:", round(p100_var, 1)])
        ws[f"B{ws.max_row}"].number_format = '0.0"%"'
        
        # --- INJECT WARNING INTO COLUMN C ---
        ws.append(["Outlier Impact Ratio (Spike Intensity) %:", round(outlier_impact, 1), oir_warning])
        ws[f"B{ws.max_row}"].number_format = '0.0"%"'
        
        # --- NEW: Append the Sensor ID string ---
        ws.append(["Absolute Worst Spot Detected:", f"{max_worst_de:.2f} ΔE (Occurred at {max_worst_dist:.3f} Yards) by Sensor {max_worst_sensor_id}"])
        ws.append([])

        # --- INJECT THE OIR LEGEND TABLE ---
        ws.append(["--- OIR (OUTLIER IMPACT RATIO) LEGEND ---"])
        r2 = ws.max_row
        ws[f"A{r2}"].font = Font(bold=True, color="00529B")
        ws.merge_cells(f"A{r2}:B{r2}")
        ws.append(["0% - 2%", "Highly Consistent: P100 and P99 are nearly identical. No sudden oil stains or color drops."])
        ws.append(["2.1% - 5%", "Moderate Variance: A few noticeable but non-critical spikes exist on the roll."])
        ws.append(["> 5%", "Severe Defect Alert: The base fabric is fine, but a massive, severe stain is ruining the average!"])
        ws.append([])

        # --- INJECT THE SQI GRADING LEGEND TABLE ---
        ws.append(["--- SQI VARIATION & GRADING LEGEND ---"])
        r_sqi = ws.max_row
        ws[f"A{r_sqi}"].font = Font(bold=True, color="00529B")
        ws.merge_cells(f"A{r_sqi}:B{r_sqi}")
        ws.append(["Grade A (<= 6.0%)", "Premium Quality: Excellent shade consistency. Meets strictest buyer tolerances."])
        ws.append(["Grade B (6.1% - 11.0%)", "Standard Quality: Acceptable commercial match. Noticeable but within passable limits."])
        ws.append(["Grade C (> 11.0%)", "Rejected: Severe shade band or heavy staining. Fails quality assurance."])
        ws.append([])

        ws.append(["--- INDIVIDUAL SENSOR YIELD DASHBOARD ---"])
        r3 = ws.max_row
        ws[f"A{r3}"].font = Font(bold=True, color="00529B")
        ws.merge_cells(f"A{r3}:L{r3}") # Expanded merge to cover all 12 columns
        
        # Dynamically change the Excel headers based on the mode!
        # --- PREMIUM FIX: Separate Master/Standard variables from Manual QTX variables ---
        if self.calib_mode == 'MANUAL':
            # Uses the "1" notation specifically for Manual QTX mode
            col_L, col_a, col_b, col_C, col_h = "Manual L1", "Manual a1", "Manual b1", "Manual C1*", "Manual h1°"
        else:
            # All Master modes (Normal & Selective) use standard spectrophotometer notation
            col_L, col_a, col_b, col_C, col_h = "Master L*", "Master a*", "Master b*", "Master C*", "Master h°"
        
        # --- NEW: Inject C1* (Chroma) and h1° (Hue) into the headers ---
        ws.append([
            "Sensor Node", "Status", 
            col_L, col_a, col_b, col_C, col_h, 
            f"Grade A (<={de_a:.2f}) %", f"Grade B (<={de_b:.2f}) %", f"Grade C (>{de_b:.2f}) %", 
            "Max Spike (ΔE CMC)", "Spike Distance (Yards)"
        ])
        
        dash_header_row = ws.max_row
        for col_idx in range(1, 13): # Extended range to 13 to cover the 2 new columns
            ws.cell(row=dash_header_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=dash_header_row, column=col_idx).alignment = Alignment(horizontal="center")
            
        for i in range(5):
            m_L = self.vault_data[i].get('L', 0.0)
            m_a = self.vault_data[i].get('a', 0.0)
            m_b = self.vault_data[i].get('b', 0.0)
            status = "ACTIVE" if self.active_sensors_state[i] else "DISABLED"
            
            # --- SPECTROPHOTOMETER MATH: Calculate Chroma (C*) and Hue (h°) ---
            if self.active_sensors_state[i]:
                # OPTIMIZATION: Direct multiplication replaces heavy exponent calculations
                m_C = math.sqrt((m_a * m_a) + (m_b * m_b))
                m_h_deg = math.degrees(math.atan2(m_b, m_a))
                if m_h_deg < 0:
                    m_h_deg += 360.0
            else:
                m_C = 0.0
                m_h_deg = 0.0
            
            # --- CRITICAL FIX: Restore pure numeric variables for downstream graph math ---
            val_L = round(m_L, 2) if self.active_sensors_state[i] else "N/A"
            val_a = round(m_a, 2) if self.active_sensors_state[i] else "N/A"
            val_b = round(m_b, 2) if self.active_sensors_state[i] else "N/A"
            val_C = round(m_C, 2) if self.active_sensors_state[i] else "N/A"
            val_h = round(m_h_deg, 2) if self.active_sensors_state[i] else "N/A"
            
            # Append the full 12-item row
            ws.append([
                f"Sensor {i}", status, 
                val_L, val_a, val_b, val_C, val_h, 
                sensor_stats[i]['A'], sensor_stats[i]['B'], sensor_stats[i]['C'], 
                sensor_stats[i]['Max'], sensor_stats[i]['MaxDist']
            ])
            for col_idx in range(2, 13):
                ws.cell(row=ws.max_row, column=col_idx).alignment = Alignment(horizontal="center")
        ws.append([])

        # =========================================================================
        # --- FORMATTING ENGINE FOR SHEET 1 (SUMMARY DASHBOARD) ---
        # =========================================================================
        center_alignment = Alignment(horizontal="center", vertical="center")
        font_purple = Font(color="7030A0")
        font_purple_bold = Font(bold=True, color="7030A0")
        font_orange_bold = Font(bold=True, color="FF8C00")
        
        fill_light_blue = PatternFill(start_color="33CCFF", end_color="33CCFF", fill_type="solid")
        fill_title = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid") 
        
        header_rows_ws1 = [dash_header_row]
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if isinstance(val, str) and val.startswith("---"):
                header_rows_ws1.append(r)
        
        for col in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                if cell.row == 1: cell.fill = fill_title
                elif cell.row in header_rows_ws1: cell.fill = fill_light_blue

                if type(cell).__name__ != 'MergedCell':
                    if col_letter is None: col_letter = cell.column_letter
                    if cell.value is not None:
                        if cell.row > 1: max_length = max(max_length, len(str(cell.value)))
                        cell.alignment = center_alignment
                        
                        if cell.row == 1:
                            cell.font = Font(bold=True, size=14, color="7030A0")
                        else:
                            is_orange = False
                            if cell.column == 1 and isinstance(cell.value, str):
                                text = cell.value
                                if text.endswith(":") or text.startswith("0%") or text.startswith("2.1%") or text.startswith("> 5%") or text.startswith("Grade A (") or text.startswith("Grade B (") or text.startswith("Grade C ("):
                                    is_orange = True

                            if is_orange: cell.font = font_orange_bold
                            else: cell.font = font_purple_bold if cell.font and cell.font.bold else font_purple
            if col_letter and max_length > 0:
                ws.column_dimensions[col_letter].width = max_length + 2

        # =========================================================================
        # --- NEW: Create Sheet 2 for Observation Data (Smart Interval Blocks) ---
        # =========================================================================
        ws2 = wb.create_sheet("Observation Data")
        
        from openpyxl.styles import Border, Side
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        next_interval_target = 0.0
        was_failing = False
        
        # --- NEW: Trackers for the Worst ΔE in the current interval ---
        interval_worst_de = {0: 0.0, 1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
        
        # OPTIMIZATION: Pre-format the constant limit string so it doesn't process 5000 times
        de_a_str = f"{de_a:.2f}"

        # --- LEVEL-3 CACHE: Pre-calculate Master Vectors before the loop ---
        master_cache = {}
        for s in range(5):
            if self.active_sensors_state[s]:
                m_L_cache = self.vault_data[s].get('L', 0.0)
                m_a_cache = self.vault_data[s].get('a', 0.0)
                m_b_cache = self.vault_data[s].get('b', 0.0)
                m_C_cache = math.sqrt((m_a_cache * m_a_cache) + (m_b_cache * m_b_cache))
                
                # OPTIMIZATION: Calculate Master Hue Angle exactly once before the massive loop starts
                m_h_deg_cache = math.degrees(math.atan2(m_b_cache, m_a_cache))
                if m_h_deg_cache < 0: m_h_deg_cache += 360.0
                
                master_cache[s] = {
                    'm_a': m_a_cache, 
                    'm_b': m_b_cache, 
                    'm_C': m_C_cache,
                    'm_h_deg': m_h_deg_cache, # Safely locked in RAM
                    'u_a': (m_a_cache / m_C_cache) if m_C_cache > 0 else 0.0,
                    'u_b': (m_b_cache / m_C_cache) if m_C_cache > 0 else 0.0
                }

        # --- OPTIMIZATION FIX: Create the massive HD Figure object ONCE in RAM ---
        report_fig = Figure(figsize=(5.0, 2.5), dpi=300)
        report_canvas = FigureCanvasAgg(report_fig)

        for i in range(total_points):
            current_dist = dists[i]
            s_dE_vals = [s_data_list[0][i], s_data_list[1][i], s_data_list[2][i], s_data_list[3][i], s_data_list[4][i]]
            
            # --- NEW: Continuously update the worst ΔE for this specific interval ---
            for s in range(5):
                if self.active_sensors_state[s] and s_dE_vals[s] > interval_worst_de[s]:
                    interval_worst_de[s] = s_dE_vals[s]
            
            currently_failing = any(self.active_sensors_state[s] and s_dE_vals[s] >= de_a for s in range(5))
            trigger_print = False
            
            # RULE 1: Standard 2-Yard Check-in (0, 2, 4, 6...)
            if current_dist >= next_interval_target:
                trigger_print = True
                next_interval_target = (math.floor(current_dist / 2.0) + 1) * 2.0
                
            # RULE 2: Instant Spike Trigger (The moment fabric enters Failure Tolerance)
            if currently_failing and not was_failing:
                trigger_print = True
                
            was_failing = currently_failing
            
            if trigger_print:
                # Add just 1 standard spacer between blocks so grids don't touch
                if ws2.max_row > 1:
                    ws2.append([])

                ws2.append([f"{current_dist:.3f} Yards"])
                start_row = ws2.max_row
                
                # 1. Build the Top Orange Distance Header
                ws2.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
                title_cell = ws2.cell(row=start_row, column=1)
                title_cell.alignment = center_alignment
                title_cell.font = Font(bold=True, color="FFFFFF", size=16) # Scaled up to 16
                title_cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
                
                # 2. Build the Purple Sensor ID Sub-Headers
                sensor_row = start_row + 1
                row_sensors = []
                for s in range(5):
                    row_sensors.extend([f"Sensor #{s}", ""])
                ws2.append(row_sensors)
                
                for s in range(5):
                    col_start = s * 2 + 1
                    ws2.merge_cells(start_row=sensor_row, start_column=col_start, end_row=sensor_row, end_column=col_start+1)
                    s_cell = ws2.cell(row=sensor_row, column=col_start)
                    s_cell.alignment = center_alignment
                    s_cell.font = Font(bold=True, color="FFFFFF", size=14) # Scaled up to 14
                    s_cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
                
                # 3. Process the Mathematical Data Matrix (Pure Floats)
                dL_row = []; da_row = []; db_row = []; dC_row = []; dH_row = []; dE_row = []; var_row = []; pf_row = []
                dH_letters = [] # Store dynamic Hue letters for formatting
                
                for s in range(5):
                    if self.active_sensors_state[s]:
                        # INSTANT RAM PULL: Bypasses redundant math
                        m_a = master_cache[s]['m_a']
                        m_b = master_cache[s]['m_b']
                        m_C = master_cache[s]['m_C']
                        
                        s_dL = det_data[s]['dL'][i] if i < len(det_data[s]['dL']) else 0.0
                        s_da = det_data[s]['da'][i] if i < len(det_data[s]['da']) else 0.0
                        s_db = det_data[s]['db'][i] if i < len(det_data[s]['db']) else 0.0
                        s_dE = s_dE_vals[s]
                        
                        # --- FIX: Calculate True SQI Variation using the WORST ΔE in this interval ---
                        worst_interval_dE = interval_worst_de[s]
                        s_ratio = worst_interval_dE / current_sigma
                        s_sqi = 100.0 * math.exp(-0.5 * (s_ratio * s_ratio))
                        s_var = 100.0 - s_sqi
                        
                        samp_a = m_a + s_da
                        samp_b = m_b + s_db
                        samp_C = math.sqrt((samp_a * samp_a) + (samp_b * samp_b))
                        
                        s_dC = samp_C - m_C
                        # THE RAW HUE FIX: Calculate physical Hue directly from physical Cartesian coordinates. 
                        # We cannot use s_dE here because it is now the weighted CMC score!
                        dH_sq = max(0.0, ((s_da * s_da) + (s_db * s_db)) - (s_dC * s_dC))
                        s_dH = math.sqrt(dH_sq)
                        
                        # --- THE TRUE QUADRANT-BASED HUE LOGIC ---
                        if m_C > 0.5:
                            cross_prod = (m_a * samp_b) - (m_b * samp_a)
                            sign_dH = 1.0 if cross_prod >= 0 else -1.0
                            s_dH_signed = s_dH * sign_dH
                            
                            # INSTANT RAM PULL: Zero calculation, zero 'if' checks
                            m_h_deg = master_cache[s]['m_h_deg']
                            
                            # Determine Primary Axis of rotation based on Master Quadrant
                            if sign_dH > 0: # Counter-Clockwise Rotation
                                if 0 <= m_h_deg < 90: h_let = "Y"      # Red -> Yellower  
                                elif 90 <= m_h_deg < 180: h_let = "G"  # Yellow -> Greener
                                elif 180 <= m_h_deg < 270: h_let = "B" # Green -> Bluer
                                else: h_let = "R"                      # Blue -> Redder
                            else: # Clockwise Rotation
                                if 0 <= m_h_deg <= 90: h_let = "R"     # Yellow -> Redder
                                elif 90 < m_h_deg <= 180: h_let = "Y"  # Green -> Yellower
                                elif 180 < m_h_deg <= 270: h_let = "G" # Blue -> Greener
                                else: h_let = "B"                      # Red -> Bluer
                        else:
                            s_dH_signed = 0.0
                            h_let = ""

                        # --- FIX: Pass/Fail must now also check the worst ΔE in the interval ---
                        pf_status = "FAIL" if worst_interval_dE >= de_a else "PASS"
                        
                        # Append raw floats to eliminate Excel text errors
                        dL_row.extend(["dL", round(s_dL, 2)])
                        da_row.extend(["da", round(s_da, 2)])
                        db_row.extend(["db", round(s_db, 2)])
                        dC_row.extend(["dC", round(s_dC, 2)])
                        dH_row.extend(["dH", round(s_dH_signed, 2)])
                        
                        # --- FIX: Display the tracked worst ΔE CMC in the grid cell ---
                        dE_row.extend(["ΔE CMC", round(worst_interval_dE, 2)])
                        
                        # --- NEW: Inject Variation as a pure float below the ΔE row ---
                        var_row.extend(["Variation", round(s_var, 2)]) 
                        
                        pf_row.extend(["P/F ΔE CMC", f"{de_a_str} ({pf_status})"])
                        dH_letters.append(h_let)
                    else:
                        dL_row.extend(["dL", "OFF"])
                        da_row.extend(["da", "OFF"])
                        db_row.extend(["db", "OFF"])
                        dC_row.extend(["dC", "OFF"])
                        dH_row.extend(["dH", "OFF"])
                        dE_row.extend(["ΔE", "OFF"])
                        
                        var_row.extend(["Variation", "OFF"])
                        
                        pf_row.extend(["P/F ΔE", "OFF"])
                        dH_letters.append("")
                
                ws2.append(dL_row)
                ws2.append(da_row)
                ws2.append(db_row)
                ws2.append(dC_row)
                ws2.append(dH_row)
                ws2.append(dE_row)
                
                # Stack the Variation row right before the Pass/Fail row
                ws2.append(var_row) 
                ws2.append(pf_row)
                
                # 4. Apply Crisp Grid Borders, Custom Number Formats, and HD Fonts
                for r in range(start_row, start_row + 10): # MATH FIX: Expanded to 10 rows
                    for c in range(1, 11):
                        cell = ws2.cell(row=r, column=c)
                        cell.border = thin_border
                        cell.alignment = center_alignment
                        
                        if r >= start_row + 2:
                            if c % 2 != 0: # Left-side Labels (dL, da, db...)
                                cell.font = Font(bold=True, color="7030A0", size=13) 
                            else: # Right-side Values
                                val = cell.value
                                s_idx = (c // 2) - 1
                                
                                if val != "OFF":
                                    cell.font = Font(bold=True, size=13) 
                                    
                                    if r == start_row + 2: # dL
                                        cell.number_format = '0.00" L";-0.00" D";0.00'
                                    elif r == start_row + 3: # da (Color-Coded)
                                        cell.number_format = '[Red]0.00" R";[Color10]-0.00" G";0.00'
                                    elif r == start_row + 4: # db (Color-Coded)
                                        cell.number_format = '[Color44]0.00" Y";[Blue]-0.00" B";0.00'
                                    elif r == start_row + 5: # dC
                                        cell.number_format = '0.00" B";-0.00" D";0.00'
                                    elif r == start_row + 6: # dH
                                        let = dH_letters[s_idx]
                                        if let == "R":
                                            cell.number_format = '[Red]0.00" R";[Red]-0.00" R";0.00'
                                        elif let == "G":
                                            cell.number_format = '[Color10]0.00" G";[Color10]-0.00" G";0.00'
                                        elif let == "Y":
                                            cell.number_format = '[Color44]0.00" Y";[Color44]-0.00" Y";0.00'
                                        elif let == "B":
                                            cell.number_format = '[Blue]0.00" B";[Blue]-0.00" B";0.00'
                                    elif r == start_row + 7: # dE
                                        cell.number_format = '0.00'
                                        if float(val) >= de_a:
                                            cell.font = Font(bold=True, color="FF0000", size=13)
                                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                            
                                    # --- NEW: Format Variation natively as a percentage ---
                                    elif r == start_row + 8: 
                                        cell.number_format = '0.00"%"'
                                        
                                    # --- SHIFTED: Pass/Fail String moved down 1 index ---
                                    elif r == start_row + 9: 
                                        if "FAIL" in str(val):
                                            cell.font = Font(bold=True, color="FF0000", size=13)
                                            
                # =========================================================================
                # THE TRUE DATACOLOR VISUAL ENGINE (DYNAMIC CMC LIMITS)
                # =========================================================================
                
                # --- OPTIMIZATION FIX: Instantly clear and reuse the pre-built RAM object ---
                report_fig.clf() 
                report_fig.subplots_adjust(top=0.82, bottom=0.15, left=0.05, right=0.95)

                gs = gridspec.GridSpec(1, 2, width_ratios=[3.5, 1.5], wspace=0.35)
                ax_wheel = report_fig.add_subplot(gs[0])
                ax_bar = report_fig.add_subplot(gs[1])

                # --- 1. CALCULATE CMC WEIGHTS (LEVEL-3 CACHE OPTIMIZED) ---
                anchor_s = next((s for s in range(5) if self.active_sensors_state[s]), 0)
                m_L = self.vault_data[anchor_s].get('L', 0.0)
                
                # INSTANT RAM PULL: Grab pre-calculated Master a & b from cache
                m_a = master_cache[anchor_s]['m_a']
                m_b = master_cache[anchor_s]['m_b']
                
                # Pull Master Hue Angle from RAM if it exists, otherwise calculate it once
                if 'm_h_rad' not in master_cache[anchor_s]:
                    master_cache[anchor_s]['m_h_rad'] = math.atan2(m_b, m_a)
                    master_cache[anchor_s]['m_h_deg'] = math.degrees(master_cache[anchor_s]['m_h_rad'])
                    
                master_hue_rad = master_cache[anchor_s]['m_h_rad']
                master_hue_deg = master_cache[anchor_s]['m_h_deg']
                
                # Extract the crucial geometric multipliers (Bypasses heavy math due to Level-2 cache)
                _, S_C, S_H, S_L = self.calc_cmc(m_L, m_a, m_b, 0.0, 0.0, 0.0)

                # Find max drift to scale the graphs safely
                max_val = 0.5
                max_dL = 0.5
                active_points = []
                for s in range(5):
                    if self.active_sensors_state[s]:
                        s_dL = det_data[s]['dL'][i] if i < len(det_data[s]['dL']) else 0.0
                        s_da = det_data[s]['da'][i] if i < len(det_data[s]['da']) else 0.0
                        s_db = det_data[s]['db'][i] if i < len(det_data[s]['db']) else 0.0
                        max_val = max(max_val, abs(s_da), abs(s_db))
                        max_dL = max(max_dL, abs(s_dL))
                        active_points.append((s, s_dL, s_da, s_db))

                # Scale the wheel safely based on the massive Chroma weight
                cmc_radius = de_a * S_C
                limit = math.ceil(max_val * 1.3 * 10) / 10.0
                if limit < cmc_radius * 1.5: limit = cmc_radius * 1.5

                limit_L = math.ceil(max_dL * 1.3 * 10) / 10.0
                l_bound = de_a * (2.0 * S_L) # The true Lightness limit
                if limit_L < l_bound * 1.5: limit_L = l_bound * 1.5

                # -------------------------------------------------------------------------
                # PART A: THE COLOR WHEEL (ax_wheel)
                # -------------------------------------------------------------------------
                ax_wheel.set_xlim(-limit, limit)
                ax_wheel.set_ylim(-limit, limit)
                ax_wheel.set_aspect('equal', adjustable='box')

                # DIAGONAL COLOR QUADRANTS
                poly_yellow = Polygon([(-limit, limit), (0,0), (limit, limit)], color='#FFFCA6', alpha=0.5)
                poly_blue   = Polygon([(-limit, -limit), (0,0), (limit, -limit)], color='#E6E6FF', alpha=0.6)
                poly_red    = Polygon([(limit, limit), (0,0), (limit, -limit)], color='#FFE6E6', alpha=0.5)
                poly_green  = Polygon([(-limit, limit), (0,0), (-limit, -limit)], color='#E6FFE6', alpha=0.5)
                ax_wheel.add_patch(poly_yellow)
                ax_wheel.add_patch(poly_blue)
                ax_wheel.add_patch(poly_red)
                ax_wheel.add_patch(poly_green)

                ax_wheel.axhline(0, color='black', linewidth=1.0, alpha=0.6)
                ax_wheel.axvline(0, color='black', linewidth=1.0, alpha=0.6)

                # DYNAMIC TICK MARKS
                num_ticks_w = 7
                step_w = (limit * 2) / (num_ticks_w - 1)
                tick_len = limit * 0.03
                for tick_idx in range(num_ticks_w):
                    t_val = -limit + (tick_idx * step_w)
                    ax_wheel.plot([t_val, t_val], [-tick_len, tick_len], color='black', lw=0.8, alpha=0.6)
                    ax_wheel.plot([-tick_len, tick_len], [t_val, t_val], color='black', lw=0.8, alpha=0.6)
                    if abs(t_val) > 0.01:
                        ax_wheel.text(t_val, -limit*0.06, f"{t_val:.1f}", fontsize=5, ha='center', va='top', alpha=0.8)
                        ax_wheel.text(-limit*0.06, t_val, f"{t_val:.1f}", fontsize=5, ha='right', va='center', alpha=0.8)

                # MASTER HUE VECTOR AND PERPENDICULAR AXIS (Exactly matching PINK.pdf bounds)
                r_C = de_a * S_C
                end_x = math.cos(master_hue_rad) * r_C 
                end_y = math.sin(master_hue_rad) * r_C
                
                r_H = de_a * S_H
                # OPTIMIZATION: Multiplication is vastly faster than division at the CPU level
                perp_rad = master_hue_rad + (math.pi * 0.5)
                p_x = math.cos(perp_rad) * r_H
                p_y = math.sin(perp_rad) * r_H
                
                # Draw the Hue Line crossing the ellipse
                ax_wheel.plot([-end_x, end_x], [-end_y, end_y], color='black', linestyle='--', linewidth=1.0, alpha=0.7)
                # Draw the Arrowhead exactly at the edge of the ellipse pointing outwards
                ax_wheel.plot(end_x, end_y, marker=(3, 0, master_hue_deg - 90), markersize=6, color='black', alpha=0.8)
                
                # Draw the Perpendicular Minor Axis Line crossing the ellipse
                ax_wheel.plot([-p_x, p_x], [-p_y, p_y], color='black', linestyle='--', linewidth=1.0, alpha=0.7)

                # --- THE TRUE CMC ELLIPSE AREA ---
                e_width = (de_a * S_C) * 2
                e_height = (de_a * S_H) * 2
                inner_ellipse = Ellipse((0,0), width=e_width, height=e_height, angle=master_hue_deg,
                                        edgecolor='#FF0000', facecolor='none', linestyle='-', linewidth=1.2, alpha=0.8)
                outer_ellipse = Ellipse((0,0), width=e_width * 1.2, height=e_height * 1.2, angle=master_hue_deg,
                                        edgecolor='gray', facecolor='none', linestyle='--', linewidth=0.8, alpha=0.7)
                ax_wheel.add_patch(inner_ellipse)
                ax_wheel.add_patch(outer_ellipse)

                ax_wheel.text(limit*0.95, 0.05 * limit, '+a*', fontsize=7, ha='right')
                ax_wheel.text(-limit*0.95, 0.05 * limit, '-a*', fontsize=7, ha='left')
                ax_wheel.text(0.05 * limit, limit*0.85, '+b*', fontsize=7, ha='left')
                ax_wheel.text(0.05 * limit, -limit*0.9, '-b*', fontsize=7, ha='left')
                ax_wheel.plot(0, 0, marker='+', color='black', markersize=8, mew=1.2)

                # -------------------------------------------------------------------------
                # PART B: THE LIGHTNESS BAR (ax_bar)
                # -------------------------------------------------------------------------
                ax_bar.set_xlim(-0.6, 0.6)
                ax_bar.set_ylim(-limit_L, limit_L)
                margin = l_bound * 0.1 

                ax_bar.add_patch(Rectangle((-0.25, -l_bound), 0.5, l_bound * 2, color='#A8E6A8', alpha=0.8)) 
                ax_bar.add_patch(Rectangle((-0.25, l_bound), 0.5, margin, color='#FFFFA6', alpha=0.9)) 
                ax_bar.add_patch(Rectangle((-0.25, -l_bound - margin), 0.5, margin, color='#FFFFA6', alpha=0.9)) 

                ax_bar.axvline(0, color='black', linewidth=1.2, alpha=0.4)
                ax_bar.axhline(0, color='black', linewidth=1.5)
                ax_bar.plot(0, 0, marker='+', color='black', markersize=8, mew=1.2)

                num_ticks = 9
                step = (limit_L * 2) / (num_ticks - 1)
                for tick_idx in range(num_ticks):
                    t_val = -limit_L + (tick_idx * step)
                    ax_bar.axhline(t_val, color='black', linewidth=0.8, alpha=0.5, xmin=0.35, xmax=0.65)
                    if abs(t_val) > 0.01: 
                        ax_bar.text(-0.35, t_val, f"{t_val:.1f}", fontsize=6, va='center', ha='right')

                ax_bar.text(0, limit_L * 1.05, '+L* (Lighter)', fontsize=7, ha='center', fontweight='bold')
                ax_bar.text(0, -limit_L * 1.15, '-L* (Darker)', fontsize=7, ha='center', fontweight='bold')

                # -------------------------------------------------------------------------
                # PART C: PLOT SENSORS
                # -------------------------------------------------------------------------
                for s_idx, s_dL, s_da, s_db in active_points:
                    color = SENSOR_COLORS[s_idx]
                    ax_wheel.plot(s_da, s_db, marker='o', color=color, markersize=4, markeredgecolor='black', mew=0.5, label=f"S{s_idx}")
                    ax_bar.plot(0, s_dL, marker='o', color=color, markersize=4, markeredgecolor='black', mew=0.5)

                # Ultra-compact legend to prevent blocking the visual data
                ax_wheel.legend(loc='upper left', fontsize=4.5, markerscale=0.6, handletextpad=0.2, borderpad=0.2, labelspacing=0.2, framealpha=0.6, edgecolor='none')

                ax_wheel.set_title(f"CIELAB Color Shift Map (Limit: {de_a:.2f} ΔE CMC)", fontsize=9, fontweight='bold', pad=6)
                ax_wheel.axis('off')
                ax_bar.axis('off')

                # Save directly to High-Speed RAM buffer
                buf = io.BytesIO()
                # --- OPTIMIZATION FIX: Use the new RAM-cached report_fig instead of fig ---
                report_fig.savefig(buf, format='png', bbox_inches='tight', pad_inches=0.1, transparent=False, facecolor='white')
                buf.seek(0)
                # (Notice we removed fig.clf() here because we now clear it at the START of the loop instead!) 

                # Stamp the image into Excel
                img = XLImage(buf)
                # --- FIX: True conversion from 295 Excel Points to 393 Screen Pixels ---
                img.width = 786
                img.height = 393 
                img.anchor = f"L{start_row}"
                ws2.add_image(img)
                
                # EXACT ROW STRETCHING
                for r_idx in range(start_row, start_row + 10): # MATH FIX: Expanded to 10 rows
                    ws2.row_dimensions[r_idx].height = 29.5
                    
                # --- NEW: Reset the interval worst trackers for the next 2-yard block ---
                interval_worst_de = {0: 0.0, 1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}

        # Set a fixed, clean width for all cells in the Observation Grid (Crash-Proofed)
        for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws2.column_dimensions[letter].width = 13

        # --- FIX 2: Catch PermissionError if the user has the Excel file open ---
        try:
            wb.save(filename)
        except PermissionError:
            self.qtx_prompt_label.configure(text=" ⚠️ ERROR: Close the Excel file first! ")
            self.after(4000, lambda: self.qtx_prompt_label.configure(text=" [ Standby ] "))
            return
        
        # --- FIX: Show success message on the actual GUI screen ---
        short_name = filename.split('/')[-1]
        self.qtx_prompt_label.configure(text=f" ✅ Excel Saved: {short_name} ")
        self.after(4000, lambda: self.qtx_prompt_label.configure(text=" [ Standby ] "))

    # ========================================================================
    # 6. HIGH-SPEED SERIAL BATCH READER THREAD
    # ========================================================================
    def serial_worker_thread(self):
        buffer = ""
        while self.is_running:
            if self.ser is None or not self.ser.is_open:
                try:
                    # 1. Save the exact distance where the wire disconnected
                    if self.last_known_distance > 0:
                        self.splice_offset_yards = self.last_known_distance
                        
                    # 2. Reconnect to ESP32
                    self.ser = serial.Serial(COM_PORT, BAUD_RATE, timeout=0.1)
                    time.sleep(2) 
                    
                    # 3. THE INVISIBLE AUTO-RESUME MACRO
                    # Only trigger if we are actively recovering from a crash (distance > 0)
                    if self.splice_offset_yards > 0:
                        self.after(0, lambda: self.qtx_prompt_label.configure(text=" 🔌 Wire Reconnected! Auto-Resuming... "))
                        
                        self.send_command('2') # Selective Mode
                        time.sleep(0.5) # Prevent USB Batching
                        for i in range(5):
                            self.send_command('1' if self.active_sensors_state[i] else '0')
                            time.sleep(0.05)
                            
                        time.sleep(0.5)
                        self.send_command('2') # Graph Mode
                        time.sleep(0.5)
                        self.send_command('3') # Manual QTX Entry
                        
                        time.sleep(1.0) # CRITICAL WAIT: Let ESP32 safely lock into calibration mode!
                        
                        # Inject the exact calibration data from the Vault
                        for i in range(5):
                            if self.active_sensors_state[i]:
                                self.send_command(str(self.vault_data[i]['L']))
                                time.sleep(0.1)
                                self.send_command(str(self.vault_data[i]['a']))
                                time.sleep(0.1)
                                self.send_command(str(self.vault_data[i]['b']))
                                time.sleep(0.1)
                                
                        time.sleep(0.5)
                        self.send_command('1') # Resume Scanning!
                        self.is_paused = False
                        self.after(0, lambda: self.qtx_prompt_label.configure(text=" [ Standby ] "))
                except Exception:
                    time.sleep(1) 
                    continue

            try:
                bytes_to_read = self.ser.in_waiting
                if bytes_to_read > 0:
                    chunk = self.ser.read(bytes_to_read)
                    buffer += chunk.decode('utf-8', errors='ignore')

                    

                    if '\n' in buffer:
                        lines = buffer.split('\n')
                        buffer = lines.pop() 
                        
                        for line in lines:
                            line = line.strip()
                            if not line:
                                continue
                                

                            # Manual QTX Terminal Prompt Capture
                            if "Target" in line or "Enter" in line or ("Sensor" in line and "Complete" in line):
                                self.pending_qtx_prompt = line
                                
                            elif "Saved" in line and "Master" not in line and ("L1:" in line or "a1:" in line or "b1:" in line):
                                self.pending_saved_message = line
                                
                                try:
                                    # Perfectly extract Manual QTX values and save them to the memory vault!
                                    match_val = re.search(r"Saved\s+([Lab])1:\s*([\-\d\.]+)", line)
                                    match_sensor = re.search(r"Sensor\s+(\d+)", line)
                                    if match_val and match_sensor:
                                        key = match_val.group(1)
                                        val_num = float(match_val.group(2))
                                        s_idx = int(match_sensor.group(1))
                                        self.vault_data[s_idx][key] = val_num
                                except Exception:
                                    pass
                                
                            elif "7-SECOND HOLD DETECTED!" in line:
                                self.pending_hard_reset = True
                            elif "3-SECOND HOLD DETECTED!" in line:
                                self.pending_soft_reset = True

                            elif line.startswith("DATA"):
                                parts = line.split(',')
                                # STREAMLINED PAYLOAD: 1 header + 1 dist + (5 sensors * 3 coords) = 17 minimum items
                                if len(parts) >= 17:
                                    try:
                                        raw_dist = float(parts[1])
                                        dist_yards = raw_dist + self.splice_offset_yards
                                        self.last_known_distance = dist_yards 
                                        
                                        deltas = []
                                        parsed_details = {0:{}, 1:{}, 2:{}, 3:{}, 4:{}}
                                        active_list = self.active_sensor_indices # INSTANT RAM PULL: Zero CPU load!

                                        for i in range(5):
                                            # HIGH-SPEED INDEXING: We now jump by 3 instead of 4
                                            base_idx = 2 + (i * 3) 
                                            raw_dL = float(parts[base_idx])
                                            raw_da = float(parts[base_idx+1])
                                            raw_db = float(parts[base_idx+2])
                                            
                                            # --- THE GHOST VAULT IS DEAD ---
                                            # The ESP32 is acting surgically. Its hardware deltas are 100% physically accurate!
                                            true_dL = raw_dL
                                            true_da = raw_da
                                            true_db = raw_db
                                            # ------------------------------------------
                                            
                                            parsed_details[i]['dL'] = true_dL
                                            parsed_details[i]['da'] = true_da
                                            parsed_details[i]['db'] = true_db
                                            
                                            if self.active_sensors_state[i]:
                                                L_std = self.vault_data[i].get('L', 0.0)
                                                a_std = self.vault_data[i].get('a', 0.0)
                                                b_std = self.vault_data[i].get('b', 0.0)
                                                
                                                # Feed the TRUE Software Deltas to the CMC formula!
                                                de_cmc, _, _, _ = self.calc_cmc(L_std, a_std, b_std, true_dL, true_da, true_db)
                                                deltas.append(de_cmc)
                                            else:
                                                deltas.append(0.0)
                                        
                                        if active_list:
                                            # OPTIMIZATION 1: Removing the [ ] brackets saves memory allocation (RAM)
                                            max_delta = max(deltas[idx] for idx in active_list)
                                            worst_idx = next(idx for idx in active_list if deltas[idx] == max_delta)
                                            
                                            rms_list = [deltas[idx] for idx in active_list]
                                            if max_delta > self.current_sigma_val and len(rms_list) > 1:
                                                rms_list.remove(max_delta)
                                                
                                            # OPTIMIZATION 2: Direct multiplication strips the heavy exponent out of the loop
                                            mean_square = sum((d * d) for d in rms_list) / len(rms_list)
                                            master_rms = math.sqrt(mean_square)
                                        else:
                                            max_delta = 0.0
                                            worst_idx = 0
                                            master_rms = 0.0

                                        self.all_distances.append(dist_yards)
                                        self.view_distances.append(dist_yards)
                                        self.all_rms.append(master_rms) 
                                        self.all_worst_de.append(max_delta)
                                        self.rms_snapshot.append(master_rms) 
                                        
                                        for i in range(5):
                                            self.all_sensor_data[i].append(deltas[i])
                                            self.view_sensor_data[i].append(deltas[i])
                                            for key in ['dL', 'da', 'db']:
                                                self.all_detailed_data[i][key].append(parsed_details[i][key])

                                        self.current_max_de = max_delta
                                        self.current_worst_sensor = worst_idx 

                                    except ValueError:
                                        # CRASH PREVENTED: USB sent a corrupted string. Ignore it and keep the thread alive!
                                        continue
                                        
                            elif "Master Saved:" in line:
                                try:
                                    # --- TRUE INDESTRUCTIBLE REGEX PARSER (FIXED) ---
                                    s_match = re.search(r'Sensor\s*[:=]?\s*(\d+)', line, re.IGNORECASE)
                                    
                                    # Added (?:1|\*)? to safely skip over '1' or '*' 
                                    # so it captures the actual data, not the letter's label!
                                    l_match = re.search(r'L(?:1|\*)?\s*[:=]?\s*([-\d.]+)', line)
                                    a_match = re.search(r'a(?:1|\*)?\s*[:=]?\s*([-\d.]+)', line)
                                    b_match = re.search(r'b(?:1|\*)?\s*[:=]?\s*([-\d.]+)', line)
                                    
                                    if not (s_match and l_match and a_match and b_match):
                                        print(f">> Parser Error: Values missing in string -> {line}")
                                        continue
                                        
                                    sensor_id = int(s_match.group(1))
                                    
                                    val_L = float(l_match.group(1))
                                    val_a = float(a_match.group(1))
                                    val_b = float(b_match.group(1))

                                    # --- THE SELECTIVE CALIBRATION SHIELD ---
                                    if getattr(self, 'calib_mode', '') == 'MASTER_SELECTIVE':
                                        if not self.allow_selective_capture:
                                            continue # Block ESP32 spam if UI button wasn't pressed
                                            
                                        if self.selective_current_idx < len(self.selective_active_list):
                                            target_s = self.selective_active_list[self.selective_current_idx]
                                            if sensor_id != target_s:
                                                continue # Ignore all other sensors reading the fabric
                                                
                                            print(f">> Selective Calibration: Sensor {target_s} successfully locked.")
                                            
                                            # --- CRITICAL FIX 2: SAVE THE VAULT BEFORE AUTO-SYNC STARTS ---
                                            self.vault_data[sensor_id]['L'] = float(l_match.group(1))
                                            self.vault_data[sensor_id]['a'] = float(a_match.group(1))
                                            self.vault_data[sensor_id]['b'] = float(b_match.group(1))
                                            # --------------------------------------------------------------
                                            
                                            self.allow_selective_capture = False
                                            self.selective_current_idx += 1
                                            
                                            # Send message safely to the GUI Queue
                                            if self.selective_current_idx < len(self.selective_active_list):
                                                next_s = self.selective_active_list[self.selective_current_idx]
                                                self.pending_saved_message = f" ✅ Sensor {target_s} Saved! Move fabric to Sensor {next_s} and click Get Data."
                                            else:
                                                self.pending_saved_message = f" ✅ Sequence Complete! Click 'Resume'."
                                                self.send_command('0') # Keep it safely paused
                                                
                                                # PROOF FOR THE TERMINAL
                                                print("\n" + "="*50)
                                                print(">> SELECTIVE CALIBRATION VAULT PRESERVED:")
                                                for s_idx in self.selective_active_list:
                                                    vd = self.vault_data.get(s_idx, {})
                                                    print(f">> Sensor {s_idx} Target -> L: {vd.get('L', 0.0)} | a: {vd.get('a', 0.0)} | b: {vd.get('b', 0.0)}")
                                                print("="*50 + "\n")
                                    # ----------------------------------------

                                    # Dynamically save the extracted values
                                    self.vault_data[sensor_id]['L'] = float(l_match.group(1))
                                    self.vault_data[sensor_id]['a'] = float(a_match.group(1))
                                    self.vault_data[sensor_id]['b'] = float(b_match.group(1))
                                    
                                except Exception as e:
                                    print(f">> ERROR parsing Master Saved: {e}")
                                    pass 
                                
                            elif "[SUCCESS] Master Calibration Complete!" in line or "[SUCCESS] New Digital Standards Saved!" in line:
                                    raw_time = datetime.now().strftime("%I:%M %p").lower()
                                    self.calibration_time_str = raw_time[1:] if raw_time.startswith("0") else raw_time
                                    self.calibration_flag = True
                                    
                                    # --- CRITICAL FIX: SEQUENCE PROTECTION ---
                                    # Prevent this generic success message from overwriting the Selective step-by-step instructions!
                                    if getattr(self, 'calib_mode', '') != 'MASTER_SELECTIVE':
                                        self.pending_saved_message = " ✅ Calibration Complete! Click 'Resume' to begin."
                                        
                                    self.send_command('0') # Always ensure the hardware stays safely paused


            except serial.SerialException:
                # CATCH DISCONNECT: Force the port closed so the Auto-Resume Macro triggers
                if self.ser:
                    try:
                        self.ser.close()
                    except:
                        pass
                self.ser = None
            time.sleep(0.005) 

    # ========================================================================
    # 7. THE GRAPHICS ENGINE 
    # ========================================================================
    def update_graphics_loop(self):
        if not self.is_running: return
        
        current_sigma = self.current_sigma_val

        with self.data_lock:
            dist_snapshot = list(self.view_distances)
            sensor_snapshot = {i: list(self.view_sensor_data[i]) for i in range(5)}
            rms_snap = list(self.rms_snapshot)
            worst_de_snap = self.current_max_de
            worst_sensor_snap = self.current_worst_sensor
            
            calib_update_needed = self.calibration_flag
            calib_time = self.calibration_time_str
            if self.calibration_flag: self.calibration_flag = False

            qtx_prompt = self.pending_qtx_prompt
            saved_msg = self.pending_saved_message
            hard_reset = self.pending_hard_reset
            soft_reset = self.pending_soft_reset

            self.pending_qtx_prompt = ""
            self.pending_saved_message = ""
            self.pending_hard_reset = False
            self.pending_soft_reset = False

        # --- FIX: Timer logic moved safely BELOW the data lock! ---
        current_time = time.time()
        dt = current_time - self.last_timer_tick
        self.last_timer_tick = current_time
        
        # 1. Grab the absolute current distance of the fabric safely
        current_dist = dist_snapshot[-1] if len(dist_snapshot) > 0 else 0.0
        
        # 2. Check if the wheel physically moved forward
        if current_dist > self.last_physical_distance:
            self.last_movement_time = current_time
            self.last_physical_distance = current_dist
            
        # 3. Wheel is considered "Spinning" if it moved in the last 0.5 seconds
        is_wheel_spinning = (current_time - self.last_movement_time) < 0.5

        # 4. Apply exact logic: Running ONLY if software is resumed AND wheel is spinning!
        if (not self.is_paused) and is_wheel_spinning:
            self.total_running_seconds += dt
        else:
            self.total_paused_seconds += dt

        # --- Live UI Update for Timers ---
        r_m, r_s = divmod(int(self.total_running_seconds), 60)
        r_h, r_m = divmod(r_m, 60)
        p_m, p_s = divmod(int(self.total_paused_seconds), 60)
        p_h, p_m = divmod(p_m, 60)
        
        run_str = f"{r_h}h {r_m}m {r_s}s" if r_h > 0 else f"{r_m}m {r_s}s"
        pause_str = f"{p_h}h {p_m}m {p_s}s" if p_h > 0 else f"{p_m}m {p_s}s"
        
        new_timer_text = f"Running: {run_str}\nPaused: {pause_str}"
        
        # STRING CACHE
        if new_timer_text != self.last_timer_text:
            self.timer_box.set_text(new_timer_text)
            self.last_timer_text = new_timer_text
            self.fig.canvas.draw_idle()

        if hard_reset:
            self._safe_trigger_hard_reset()
            return
        if soft_reset:
            self._safe_trigger_soft_reset()

        if qtx_prompt:
            self._safe_update_qtx_prompt(qtx_prompt)

        if saved_msg:
            self._safe_show_saved(saved_msg)

        if calib_update_needed:
            self.calib_box.set_text(f"Calibration done: {calib_time}")
            # --- CRITICAL FIX: Do NOT overwrite the Selective Mode success text! ---
            if getattr(self, 'calib_mode', '') != 'MASTER_SELECTIVE':
                self.qtx_prompt_label.configure(text=" [ Standby ] ")
                self.terminal_entry.delete(0, 'end')
            self.fig.canvas.draw_idle()

        if len(dist_snapshot) > 0:
            for i in range(5):
                if self.active_sensors_state[i]:
                    view_data = sensor_snapshot[i]
                    self.lines[i].set_data(dist_snapshot, view_data)
                else:
                    self.lines[i].set_data([], [])
            
            self.rms_line.set_data(dist_snapshot, rms_snap)
            
            # Update left graph X and Y limits
            x_min = dist_snapshot[0]
            x_max = dist_snapshot[-1] + 2
            self.ax1.set_xlim(x_min, x_max) 
            
            # --- FIX: Zoom in the Y-axis further to (Sigma + 0.5) ---
            self.ax1.set_ylim(0, current_sigma + 0.5) 
            
            # --- UPDATE HORIZONTAL RAINBOW LINE ON LEFT GRAPH ---
            de_a = current_sigma * SIGMA_FACTOR_A
            line_thickness = (current_sigma + 0.5) * 0.008 # Keeps thickness perfectly proportionate to the extreme zoom
            self.rainbow_line.set_extent([x_min, x_max, de_a - line_thickness, de_a + line_thickness])

            if current_sigma != self.last_sigma:
                self.sigma_line.set_ydata([current_sigma, current_sigma])
                max_x = max(5.0, current_sigma * 3.0) 
                self.ax2.set_xlim(0, max_x)
                x_vals = np.linspace(0, max_x, 200)
                
                # OPTIMIZATION: Even in Numpy, direct array multiplication is faster than the **2 power function!
                ratio_arr = x_vals / current_sigma
                y_vals = 100.0 * np.exp(-0.5 * (ratio_arr * ratio_arr))
                
                self.line_curve.set_data(x_vals, y_vals)
                self.last_sigma = current_sigma

            current_rms_val = rms_snap[-1] if len(rms_snap) > 0 else 0.0
            
            # OPTIMIZATION: math.exp() skips Numpy's heavy array-checking overhead. Direct multiplication removes **2.
            rms_ratio = current_rms_val / current_sigma
            worst_ratio = worst_de_snap / current_sigma
            
            sqi_rms = 100.0 * math.exp(-0.5 * (rms_ratio * rms_ratio))
            sqi_worst = 100.0 * math.exp(-0.5 * (worst_ratio * worst_ratio))
            
            var_rms = 100.0 - sqi_rms
            var_worst = 100.0 - sqi_worst
            
            self.dot_current.set_data([current_rms_val], [sqi_rms]) 
            self.dot_worst.set_data([worst_de_snap], [sqi_worst])   

            latest_vals_str = []
            for i in range(5):
                if self.active_sensors_state[i]:
                    val = sensor_snapshot[i][-1] if len(sensor_snapshot[i]) > 0 else 0.0
                    latest_vals_str.append(f"Sensor {i}: {val:.2f}")
                else:
                    latest_vals_str.append(f"Sensor {i}: Off")
                    
            new_live_text = f"Distance: {dist_snapshot[-1]:.3f} Yards\n-----------------\nCurrent ΔE CMC:\n" + "\n".join(latest_vals_str) + f"\nMaster RMS: {current_rms_val:.2f}"
            if new_live_text != self.last_live_text:
                self.live_data_box.set_text(new_live_text)
                self.last_live_text = new_live_text

            current_worst_str = f"{worst_de_snap:.2f} =>> Sensor {worst_sensor_snap}"
            new_stats_text = ""
            
            if not self.is_paused:
                if var_rms < self.record_min_var: self.record_min_var = var_rms
                if var_rms > self.record_max_var: self.record_max_var = var_rms
                if worst_de_snap > self.record_max_de: 
                    self.record_max_de = worst_de_snap
                    self.record_max_sensor = worst_sensor_snap
                    self.record_max_dist = dist_snapshot[-1]
                new_stats_text = f"Min Variation: {self.record_min_var:.1f}%\nMax Variation: {self.record_max_var:.1f}%\nWorst ΔE CMC = {current_worst_str}\nMax ΔE CMC = {self.record_max_de:.2f} =>> Sensor {self.record_max_sensor} =>> {self.record_max_dist:.1f} Yards"

                if new_stats_text != self.last_stats_text:
                    self.stats_box.set_text(new_stats_text)
                    self.last_stats_text = new_stats_text

            readout_string = f"RMS Var: {var_rms:.1f}% | Worst: {var_worst:.1f}%"
            if self.p99_variation_text != "": readout_string += self.p99_variation_text
                
            self.sqi_readout.configure(text=readout_string)
            self.fig.canvas.draw_idle()

        self.after(30, self.update_graphics_loop)

# ============================================================================
# 8. STAGE 1: THE CONSOLE BOOTLOADER (RESTORED MEMORY FIX)
# ============================================================================
def run_console_bootloader():
    print("\n" + "="*60)
    print(" TMDM SHADE INSPECTION - WAITING FOR HARDWARE ")
    print("="*60)
    print(">> ACTION REQUIRED: Press the physical button on the ESP32 1 time to enter Offline Mode!")

    try:
        ser = serial.Serial(COM_PORT, BAUD_RATE, timeout=0.1)
    except Exception as e:
        print(f"\n[ERROR] FAILED TO CONNECT TO {COM_PORT}.")
        print("Check your USB cable and ensure the Arduino IDE Serial Monitor is CLOSED.")
        return None, None, None

    # CRITICAL FIX: Do NOT reboot the ESP32. Let it wait in Asking Mode!
    ser.setDTR(False)
    ser.setRTS(False)

    gui_ready = threading.Event()
    active_sensors = [True, True, True, True, True]
    calib_mode = ['MASTER']
    hardware_ready = [False]

    def read_serial():
        buffer = ""
        while not gui_ready.is_set():
            try:
                if ser.in_waiting > 0:
                    raw_data = ser.read(ser.in_waiting).decode('utf-8', errors='ignore')
                    sys.stdout.write(raw_data)
                    sys.stdout.flush()
                    
                    buffer += raw_data
                    while '\n' in buffer:
                        line, buffer = buffer.split('\n', 1)
                        # When the user physically presses the button 1 time:
                        if "OFFLINE (USB) MODE SELECTED" in line:
                            hardware_ready[0] = True
                            print("\n\n>> Hardware connected! Select Calibration Mode:")
                            print(" [1] Standard Master")
                            print(" [2] Selective Sequential Master")
                            print(" [3] Manual QTX Entry")
                            print(">> Choice: ", end='', flush=True)
            except: pass
            time.sleep(0.01)

    def read_keyboard():
        while not gui_ready.is_set():
            if hardware_ready[0]:
                try:
                    cmd = sys.stdin.readline().strip()
                    if cmd == '1':
                        calib_mode[0] = 'MASTER'
                        gui_ready.set()
                    elif cmd == '2':
                        calib_mode[0] = 'MASTER_SELECTIVE'
                        gui_ready.set()
                    elif cmd == '3':
                        calib_mode[0] = 'MANUAL'
                        gui_ready.set()
                    elif cmd != '':
                        print(">> Invalid choice. Type 1, 2, or 3: ", end='', flush=True)
                except: pass
            time.sleep(0.1)

    t_ser = threading.Thread(target=read_serial, daemon=True)
    t_kb = threading.Thread(target=read_keyboard, daemon=True)
    t_ser.start()
    t_kb.start()

    gui_ready.wait()

    print("\n" + "="*60)
    print(" GRAPHICS ENGINE ACTIVATED. LAUNCHING DASHBOARD... ")
    print("="*60 + "\n")
    time.sleep(1)

    return ser, active_sensors, calib_mode[0]

if __name__ == "__main__":
    while True:
        active_serial_connection, sensors_state, mode = run_console_bootloader()
        
        if active_serial_connection:
            app = TMDMDashboard(active_serial_connection, sensors_state, mode)
            app.mainloop()
        else:
            break
